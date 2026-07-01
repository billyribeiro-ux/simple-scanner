from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl import load_workbook

from app.db.repositories import RepositoryRegistry
from app.models.replay_evidence import (
    REPLAY_AWARE_MODEL_TYPE,
    REPLAY_AWARE_VALIDATION_MODE,
    REPLAY_AWARE_VALIDATION_PURPOSE,
)
from app.schemas.market import Bar
from app.services.research import (
    ChampionChallengerComparisonService,
    ModelProposalService,
    ResearchCycleService,
    ResearchStatusService,
)
from app.services.workflows import ExportWorkflowService
from app.utils.time import UTC


def _sheet_names(path: str) -> set[str]:
    workbook = load_workbook(path, read_only=True)
    try:
        return set(workbook.sheetnames)
    finally:
        workbook.close()


def _bars() -> list[Bar]:
    start = datetime(2026, 6, 1, 13, 30, tzinfo=UTC)
    return [
        Bar(
            symbol="AAPL",
            interval="1min",
            timestamp_utc=start + timedelta(minutes=index),
            timestamp_et=start + timedelta(minutes=index),
            open=100 + index * 0.1,
            high=100.4 + index * 0.1,
            low=99.8 + index * 0.1,
            close=100.1 + index * 0.1,
            volume=1_000 + index,
            source="phase11",
        )
        for index in range(10)
    ]


def _model(version: str, *, active: bool = False, average_r: float = 0.3, profit_factor: float = 1.4) -> dict[str, object]:
    return {
        "model_version": version,
        "model_type": REPLAY_AWARE_MODEL_TYPE,
        "activation_decision": "accepted",
        "active": active,
        "metrics": {
            "average_r": average_r,
            "profit_factor": profit_factor,
            "max_drawdown_r": -1.0,
            "observed_outcome_count": 12,
            "total_trades": 12,
        },
        "validation_metrics": {},
        "created_at": datetime.now(UTC).isoformat(),
    }


def _accepted_validation(repo: RepositoryRegistry, model_version: str) -> dict[str, object]:
    return repo.validation_reports.save(
        {
            "model_version": model_version,
            "validation_mode": REPLAY_AWARE_VALIDATION_MODE,
            "summary": {"selected_candidate_count": 8, "average_r": 0.4, "profit_factor": 1.6},
            "windows": [],
            "activation_decision": "accepted",
            "rejection_reasons": [],
            "created_at": datetime.now(UTC).isoformat(),
        },
        model_version=model_version,
        purpose=REPLAY_AWARE_VALIDATION_PURPOSE,
    )


def _passing_governance_artifacts(repo: RepositoryRegistry, model_version: str) -> tuple[dict[str, object], dict[str, object], dict[str, object]]:
    validation = _accepted_validation(repo, model_version)
    calibration = repo.model_calibration_audits.save(
        {
            "calibration_audit_id": f"calibration-{model_version}",
            "model_version": model_version,
            "replay_run_ids": [],
            "outcome_source": "counterfactual_preferred",
            "score_bins": [{"bin_key": "75-100", "sample_size": 8, "observed_average_r": 0.5}],
            "grade_bins": [{"bin_key": "A", "sample_size": 8, "observed_average_r": 0.5}],
            "action_bins": [{"bin_key": "TAKE", "sample_size": 8, "observed_average_r": 0.5}],
            "rank_correlation_score": 0.8,
            "monotonicity_pass": True,
            "separation_metrics": {"take_minus_watch_average_r": 0.4},
            "stability_metrics": {},
            "calibration_warnings": [],
            "rejection_reasons": [],
        }
    )
    drift = repo.model_calibration_drift.save(
        {
            "drift_report_id": f"drift-{model_version}",
            "model_version": model_version,
            "calibration_audit_ids": [calibration["calibration_audit_id"]],
            "summary": {"severity": "INFO"},
            "severity": "INFO",
            "drift_flags": [],
            "window_metrics": [],
        }
    )
    review = repo.model_review_reports.save(
        {
            "review_report_id": f"review-{model_version}",
            "model_version": model_version,
            "validation_report_ids": [validation["report_id"]],
            "calibration_audit_ids": [calibration["calibration_audit_id"]],
            "drift_report_ids": [drift["drift_report_id"]],
            "summary": {"readiness_status": "PASS", "model_activation_unchanged": True},
            "readiness_status": "PASS",
            "readiness_reasons": [],
            "unresolved_warnings": [],
        }
    )
    return calibration, drift, review


def _mark_clean(repo: RepositoryRegistry) -> None:
    start = _bars()[0].timestamp_utc
    end = _bars()[-1].timestamp_utc
    for artifact, version in (
        ("features", "features.v2.no_leakage"),
        ("candidates", "candidate_signals.v1"),
        ("labels", "labels.v2.no_leakage"),
        ("replay", "candidate_market_replay"),
    ):
        repo.pipeline_windows.mark_built(artifact, ["AAPL"], ["1min"], start, end, version)


def test_research_cycle_blocks_stale_and_refresh_without_secret(tmp_path, monkeypatch) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "phase11.sqlite3")
    repo.bars.upsert_many(_bars())
    monkeypatch.delenv("FMP_API_KEY", raising=False)
    service = ResearchCycleService(repo)
    created = service.create(
        {
            "cycle_date": "2026-07-01",
            "symbols": ["APPL"],
            "intervals": ["1min"],
            "start": _bars()[0].timestamp_utc.isoformat(),
            "end": _bars()[-1].timestamp_utc.isoformat(),
        }
    )
    assert created["symbols"] == ["AAPL"]
    dry_run = service.dry_run(created["research_cycle_id"])
    assert dry_run["status"] == "dry_run"
    assert dry_run["blocked"] is True
    assert dry_run["block_reason"] == "stale_artifacts_present"

    blocked = service.run(created["research_cycle_id"])
    assert blocked["status"] == "blocked"
    assert blocked["failed_reason"] == "stale_artifacts_present"

    refresh = service.create(
        {
            "cycle_date": "2026-07-01",
            "symbols": ["AAPL"],
            "intervals": ["1min"],
            "start": _bars()[0].timestamp_utc.isoformat(),
            "end": _bars()[-1].timestamp_utc.isoformat(),
            "refresh_data": True,
            "allow_stale": True,
        }
    )
    refresh_block = service.run(refresh["research_cycle_id"])
    assert refresh_block["status"] == "blocked"
    assert refresh_block["failed_reason"] == "fmp_api_key_required_for_refresh_data"


def test_research_cycle_proposes_challenger_without_activation(tmp_path) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "cycle.sqlite3")
    repo.bars.upsert_many(_bars())
    _mark_clean(repo)
    champion = repo.model_runs.save(_model("champion", active=True, average_r=0.2, profit_factor=1.2))
    repo.active_models.activate(champion, validation_report_id="champion-validation")
    repo.model_runs.save(_model("challenger", average_r=0.5, profit_factor=2.0))
    _passing_governance_artifacts(repo, "challenger")

    service = ResearchCycleService(repo)
    cycle = service.create(
        {
            "cycle_date": "2026-07-01",
            "symbols": ["AAPL"],
            "intervals": ["1min"],
            "start": _bars()[0].timestamp_utc.isoformat(),
            "end": _bars()[-1].timestamp_utc.isoformat(),
            "challenger_model_version": "challenger",
        }
    )
    result = service.run(cycle["research_cycle_id"])
    assert result["status"] == "completed"
    assert result["summary"]["model_activation_unchanged"] is True
    assert result["proposal"]["status"] == "PROPOSED"
    assert result["proposal"]["recommended_action"] == "APPROVE_CHALLENGER_FOR_ACTIVATION"
    assert repo.active_models.get_active(REPLAY_AWARE_MODEL_TYPE)["model_version"] == "champion"
    assert repo.research_cycles.list_artifacts(cycle["research_cycle_id"])


def test_proposal_approval_and_explicit_activation_lifecycle(tmp_path) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "proposal.sqlite3")
    champion = repo.model_runs.save(_model("champion", active=True, average_r=0.2, profit_factor=1.1))
    repo.active_models.activate(champion, validation_report_id="champion-validation")
    repo.model_runs.save(_model("challenger", average_r=0.6, profit_factor=2.2))
    _passing_governance_artifacts(repo, "challenger")
    comparison = ChampionChallengerComparisonService(repo).compare("champion", "challenger")
    proposal = ModelProposalService(repo).create_from_comparison(comparison)
    proposal_id = proposal["proposal_id"]

    approved = ModelProposalService(repo).approve(proposal_id, actor="research_lead")
    assert approved["status"] == "APPROVED_FOR_ACTIVATION"
    assert repo.active_models.get_active(REPLAY_AWARE_MODEL_TYPE)["model_version"] == "champion"

    blocked = ModelProposalService(repo).activate(proposal_id, actor="research_lead")
    assert blocked["status"] == "blocked"
    assert blocked["reason"] == "manual_confirmation_required"
    assert repo.active_models.get_active(REPLAY_AWARE_MODEL_TYPE)["model_version"] == "champion"

    activated = ModelProposalService(repo).activate(
        proposal_id,
        actor="research_lead",
        confirm_manual_activation=True,
        validation_mode=REPLAY_AWARE_VALIDATION_MODE,
    )
    assert activated["status"] == "ok"
    assert activated["activation"]["activated"] is True
    assert repo.active_models.get_active(REPLAY_AWARE_MODEL_TYPE)["model_version"] == "challenger"
    assert {row["decision_type"] for row in repo.model_decision_ledger.list(proposal_id=proposal_id)} >= {
        "PROPOSAL_APPROVED",
        "MODEL_ACTIVATION_REQUESTED",
        "MODEL_ACTIVATED",
    }


def test_proposal_rejects_block_readiness_and_rejected_activation(tmp_path) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "reject.sqlite3")
    repo.model_runs.save(_model("challenger", average_r=0.6, profit_factor=2.2))
    _accepted_validation(repo, "challenger")
    repo.model_calibration_drift.save(
        {
            "drift_report_id": "drift-block",
            "model_version": "challenger",
            "summary": {"severity": "BLOCKING"},
            "severity": "BLOCKING",
            "drift_flags": ["rank_correlation_deteriorating"],
            "window_metrics": [],
        }
    )
    comparison = ChampionChallengerComparisonService(repo).compare(None, "challenger")
    assert comparison["readiness_status"] == "BLOCK"
    assert comparison["recommended_action"] == "REJECT_CHALLENGER"
    proposal = ModelProposalService(repo).create_from_comparison(comparison)
    assert proposal["status"] == "REJECTED"
    activation = ModelProposalService(repo).activate(
        proposal["proposal_id"],
        confirm_manual_activation=True,
        validation_mode=REPLAY_AWARE_VALIDATION_MODE,
    )
    assert activation["status"] == "blocked"
    assert activation["reason"] == "proposal_rejected"
    assert any(row["decision_type"] == "MODEL_ACTIVATION_BLOCKED" for row in repo.model_decision_ledger.list(proposal_id=proposal["proposal_id"]))


def test_keep_champion_recommendation_cannot_be_approved_for_activation(tmp_path) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "keep.sqlite3")
    champion = repo.model_runs.save(_model("champion", active=True, average_r=0.8, profit_factor=2.8))
    repo.active_models.activate(champion, validation_report_id="champion-validation")
    repo.model_runs.save(_model("challenger", average_r=0.1, profit_factor=1.0))
    _passing_governance_artifacts(repo, "challenger")

    comparison = ChampionChallengerComparisonService(repo).compare("champion", "challenger")
    assert comparison["recommended_action"] == "KEEP_CHAMPION"
    proposal = ModelProposalService(repo).create_from_comparison(comparison)
    assert proposal["status"] == "REVIEW_REQUIRED"

    approval = ModelProposalService(repo).approve(proposal["proposal_id"], actor="research_lead")
    assert approval["status"] == "blocked"
    assert approval["reason"] == "proposal_not_recommended_for_activation"
    activation = ModelProposalService(repo).activate(
        proposal["proposal_id"],
        actor="research_lead",
        confirm_manual_activation=True,
        validation_mode=REPLAY_AWARE_VALIDATION_MODE,
    )
    assert activation["status"] == "blocked"
    assert activation["reason"] == "proposal_not_recommended_for_activation"
    assert repo.active_models.get_active(REPLAY_AWARE_MODEL_TYPE)["model_version"] == "champion"


def test_status_ledger_and_exports(tmp_path) -> None:
    repo = RepositoryRegistry(db_path=tmp_path / "status.sqlite3")
    repo.bars.upsert_many(_bars())
    _mark_clean(repo)
    champion = repo.model_runs.save(_model("champion", active=True, average_r=0.2, profit_factor=1.2))
    repo.active_models.activate(champion, validation_report_id="champion-validation")
    repo.model_runs.save(_model("challenger", average_r=0.5, profit_factor=2.0))
    _passing_governance_artifacts(repo, "challenger")
    cycle = ResearchCycleService(repo).create(
        {
            "cycle_date": "2026-07-01",
            "symbols": ["AAPL"],
            "intervals": ["1min"],
            "start": _bars()[0].timestamp_utc.isoformat(),
            "end": _bars()[-1].timestamp_utc.isoformat(),
            "challenger_model_version": "challenger",
        }
    )
    result = ResearchCycleService(repo).run(cycle["research_cycle_id"])
    proposal_id = result["proposal"]["proposal_id"]
    comparison_id = result["comparison"]["comparison_id"]
    status = ResearchStatusService(repo).status()
    assert status["status"] == "ok"
    assert status["latest_research_cycle"]["research_cycle_id"] == cycle["research_cycle_id"]
    assert status["pending_proposals"]
    assert "dirty_window_count" in status["stale_windows_summary"]

    exporter = ExportWorkflowService(repo)
    exporter.exporter.settings.exports_dir = tmp_path
    cycle_xlsx = exporter.export_research_cycle(cycle["research_cycle_id"], "xlsx")
    cycle_json = exporter.export_research_cycle(cycle["research_cycle_id"], "json")
    proposal_xlsx = exporter.export_model_proposal(proposal_id, "xlsx")
    proposal_json = exporter.export_model_proposal(proposal_id, "json")
    comparison_xlsx = exporter.export_champion_challenger_comparison(comparison_id)
    assert {"Summary", "Cycle Config", "Data Quality", "Champion vs Challenger", "Proposal", "Provenance"} <= _sheet_names(cycle_xlsx["path"])
    assert {"Summary", "Recommended Action", "Readiness", "Approval History", "Provenance"} <= _sheet_names(proposal_xlsx["path"])
    assert {"Summary", "Champion", "Challenger", "Delta Metrics", "Gates"} <= _sheet_names(comparison_xlsx["path"])
    assert cycle_json["export"]["file_sha256"]
    assert proposal_json["export"]["file_sha256"]
    assert comparison_xlsx["export"]["file_sha256"]
    decisions = repo.model_decision_ledger.list(research_cycle_id=cycle["research_cycle_id"])
    assert decisions
