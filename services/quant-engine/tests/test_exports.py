import csv
import json
from datetime import datetime

from openpyxl import load_workbook

from app.db.repositories import RepositoryRegistry
from app.exports.service import ExportService
from app.schemas.market import Side, Signal
from app.utils.time import UTC


def _sheet_names(path) -> set[str]:
    workbook = load_workbook(path, read_only=True)
    try:
        return set(workbook.sheetnames)
    finally:
        workbook.close()


def test_signal_export_csv_and_xlsx(tmp_path, monkeypatch) -> None:
    service = ExportService()
    monkeypatch.setattr(service.settings, "exports_dir", tmp_path)
    signal = Signal(
        timestamp=datetime(2026, 6, 1, 14, 0, tzinfo=UTC),
        ticker="AAPL",
        side=Side.LONG,
        entry_price=100,
        stop_price=99,
        target_1=101,
        target_2=101.5,
        target_3=102.5,
        risk_per_share=1,
        reward_risk_to_t1=1,
        reward_risk_to_t2=1.5,
        reward_risk_to_t3=2.5,
        expected_r=0.2,
        confidence_score=0.75,
        signal_grade="B+",
        setup_type="VWAP reclaim long",
        market_regime="trend_long",
        ticker_regime="single_stock_momentum",
        reasons=["test"],
        warnings=[],
        historical_sample_size=50,
        historical_win_rate=0.6,
        historical_average_r=0.3,
        model_version="test",
    )
    csv_path = service.export_signals_csv([signal])
    xlsx_path = service.export_signals_xlsx([signal])
    assert csv_path.exists()
    assert xlsx_path.exists()


def test_replay_exports_include_trades_skips_metrics_and_workbook_sheets(tmp_path, monkeypatch) -> None:
    service = ExportService()
    monkeypatch.setattr(service.settings, "exports_dir", tmp_path)
    replay_run = {
        "replay_run_id": "replay-export-test",
        "simulation_type": "candidate_market_replay",
        "created_at": datetime(2026, 6, 1, 14, 0, tzinfo=UTC).isoformat(),
        "config": {"symbols": ["AAPL"], "intervals": ["1min"], "entry_mode": "next_bar_open"},
        "summary_metrics": {
            "total_candidates": 2,
            "candidates_taken": 1,
            "candidates_skipped": 1,
            "total_trades": 1,
            "expectancy_r": 1.5,
            "profit_factor": 0,
            "skip_breakdown": {"missing_entry_bar": 1},
            "per_symbol_metrics": {"AAPL": {"total_trades": 1, "expectancy_r": 1.5}},
            "per_setup_metrics": {"VWAP reclaim long": {"total_trades": 1, "expectancy_r": 1.5}},
            "per_regime_metrics": {"trend_long": {"total_trades": 1}},
            "per_time_bucket_metrics": {"opening_drive": {"total_trades": 1}},
            "daily_r_series": [{"date": "2026-06-01", "r": 1.5}],
            "drawdown_series": [0.0],
        },
        "warnings": ["test warning"],
    }
    trades = [
        {
            "trade_id": "trade-1",
            "replay_run_id": "replay-export-test",
            "candidate_id": "candidate-1",
            "symbol": "AAPL",
            "interval": "1min",
            "side": "LONG",
            "setup_type": "VWAP reclaim long",
            "signal_timestamp_utc": datetime(2026, 6, 1, 13, 30, tzinfo=UTC).isoformat(),
            "entry_timestamp_utc": datetime(2026, 6, 1, 13, 31, tzinfo=UTC).isoformat(),
            "exit_timestamp_utc": datetime(2026, 6, 1, 13, 36, tzinfo=UTC).isoformat(),
            "entry_price": 100,
            "stop_price": 99,
            "target_1": 101,
            "target_2": 101.5,
            "target_3": 102.5,
            "exit_price": 101.5,
            "exit_reason": "target_2",
            "realized_r": 1.5,
            "mfe_r": 1.7,
            "mae_r": -0.2,
            "bars_held": 5,
            "minutes_held": 5,
            "same_bar_ambiguous": False,
            "ambiguity_policy": "conservative_stop_first",
            "slippage_bps": 1,
            "spread_bps": 2,
            "commission": 0,
            "market_regime": "trend_long",
            "time_bucket": "opening_drive",
            "signal_score": 0.8,
            "expected_r": 1.5,
            "status": "TAKEN",
        },
        {
            "trade_id": "trade-2",
            "replay_run_id": "replay-export-test",
            "candidate_id": "candidate-2",
            "symbol": "AAPL",
            "interval": "1min",
            "side": "LONG",
            "setup_type": "VWAP reclaim long",
            "signal_timestamp_utc": datetime(2026, 6, 1, 15, 59, tzinfo=UTC).isoformat(),
            "status": "SKIPPED",
            "skip_reason": "missing_entry_bar",
        },
    ]

    trades_csv = service.export_replay_trades_csv("replay-export-test", trades)
    trades_xlsx = service.export_replay_trades_xlsx("replay-export-test", trades)
    summary_xlsx = service.export_replay_summary_xlsx(replay_run, trades)
    metrics_json = service.export_replay_metrics_json(replay_run)

    with trades_csv.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 2
    assert rows[0]["status"] == "TAKEN"
    assert rows[1]["skip_reason"] == "missing_entry_bar"
    assert _sheet_names(trades_xlsx) == {"Trades"}
    assert {
        "Summary",
        "Trades",
        "Skipped Candidates",
        "Per Symbol",
        "Per Setup",
        "Per Regime",
        "Per Time Bucket",
        "Daily R",
        "Drawdown",
        "Config",
        "Warnings",
    } <= _sheet_names(summary_xlsx)
    metrics_payload = json.loads(metrics_json.read_text(encoding="utf-8"))
    assert metrics_payload["simulation_type"] == "candidate_market_replay"
    assert metrics_payload["summary_metrics"]["total_candidates"] == 2


def test_sensitivity_exports_and_repository_metadata(tmp_path, monkeypatch) -> None:
    service = ExportService()
    monkeypatch.setattr(service.settings, "exports_dir", tmp_path)
    sensitivity = {
        "sensitivity_run_id": "sensitivity-export-test",
        "replay_run_id": "replay-export-test",
        "created_at": datetime(2026, 6, 1, 14, 0, tzinfo=UTC).isoformat(),
        "config": {"slippage_bps_grid": [0, 1], "spread_bps_grid": [0, 2]},
        "scenario_count": 2,
        "robustness_score": 0.5,
        "pass_fail": "fail",
        "fragility_flags": ["robustness_score_below_threshold"],
        "gate_results": {"robustness_score_met": False},
        "worst_case": {"scenario_id": "scenario-2", "summary_metrics": {"average_r": -0.1}},
        "median_case": {"scenario_id": "scenario-1", "summary_metrics": {"average_r": 0.1}},
        "best_case": {"scenario_id": "scenario-1", "summary_metrics": {"average_r": 0.1}},
    }
    scenarios = [
        {
            "scenario_id": "scenario-1",
            "sensitivity_run_id": "sensitivity-export-test",
            "replay_run_id": "replay-export-test",
            "slippage_bps": 0,
            "spread_bps": 0,
            "intrabar_path_policy": "conservative",
            "same_bar_stop_target_policy": "conservative_stop_first",
            "pass_fail": "pass",
            "summary_metrics": {"total_trades": 10, "average_r": 0.1, "profit_factor": 1.2, "max_drawdown_r": -1, "total_r": 1, "skip_rate": 0.0},
        },
        {
            "scenario_id": "scenario-2",
            "sensitivity_run_id": "sensitivity-export-test",
            "replay_run_id": "replay-export-test",
            "slippage_bps": 1,
            "spread_bps": 2,
            "intrabar_path_policy": "conservative",
            "same_bar_stop_target_policy": "conservative_stop_first",
            "pass_fail": "fail",
            "summary_metrics": {"total_trades": 10, "average_r": -0.1, "profit_factor": 0.8, "max_drawdown_r": -2, "total_r": -1, "skip_rate": 0.0},
        },
    ]

    summary_xlsx = service.export_sensitivity_summary_xlsx(sensitivity, scenarios)
    scenarios_csv = service.export_sensitivity_scenarios_csv("sensitivity-export-test", scenarios)
    scenarios_xlsx = service.export_sensitivity_scenarios_xlsx("sensitivity-export-test", scenarios)
    metrics_json = service.export_sensitivity_metrics_json(sensitivity, scenarios)

    assert {
        "Summary",
        "Scenario Metrics",
        "Worst Case",
        "Median Case",
        "Best Case",
        "Fragility Flags",
        "Gate Results",
        "Config",
        "Warnings",
    } <= _sheet_names(summary_xlsx)
    assert _sheet_names(scenarios_xlsx) == {"Scenario Metrics"}
    assert scenarios_csv.exists()
    assert json.loads(metrics_json.read_text(encoding="utf-8"))["scenario_count"] == 2

    repo = RepositoryRegistry(db_path=tmp_path / "exports.sqlite3")
    record = repo.exports.record(
        "replay_sensitivity_summary",
        "xlsx",
        summary_xlsx,
        row_count=2,
        source_run_id="replay-export-test",
        payload={"source_simulation_type": "candidate_market_replay"},
    )
    assert record["file_sha256"]
    assert "Summary" in record["workbook_sheets"]
    persisted = repo.exports.list_all()[0]
    assert persisted["file_sha256"] == record["file_sha256"]


def test_replay_aware_exports_include_required_workbook_sheets(tmp_path, monkeypatch) -> None:
    service = ExportService()
    monkeypatch.setattr(service.settings, "exports_dir", tmp_path)
    model = {
        "model_version": "replay-aware-export-test",
        "model_type": "replay_aware_baseline",
        "metrics": {"candidate_outcome_rows": 10, "observed_outcome_count": 8, "average_r": 0.4},
        "replay_run_ids": ["replay-export-test"],
        "candidate_outcome_row_count": 10,
        "scoring_config": {"minimum_observed_outcomes": 5, "minimum_profit_factor": 1.05},
        "activation_criteria": {"minimum_selected_trades": 5},
        "warnings": [],
        "replay_config_hashes": ["hash"],
        "input_fingerprints": ["input"],
        "code_version": "test",
    }
    cells = [
        {
            "model_version": "replay-aware-export-test",
            "cell_key": "side_global:side=LONG",
            "dimensions": {"side": "LONG", "symbol": "AAPL", "setup_type": "VWAP reclaim long", "market_regime": "trend_long", "time_bucket": "opening_drive"},
            "hierarchy_level": "side_global",
            "parent_cell_key": None,
            "sample_size": 8,
            "observed_outcome_count": 8,
            "average_r": 0.4,
            "median_r": 0.4,
            "profit_factor": 1.6,
            "max_drawdown_r": -1,
            "robustness_score": 1.0,
            "fragility_flags": [],
            "metrics": {"label_vs_replay_divergence_flags": []},
            "evidence_quality_grade": "B",
        }
    ]
    audits = [
        {
            "score_id": "score-export-test",
            "model_version": "replay-aware-export-test",
            "candidate_id": "candidate-export-test",
            "symbol": "AAPL",
            "interval": "1min",
            "timestamp_utc": datetime(2026, 6, 1, 14, 0, tzinfo=UTC).isoformat(),
            "side": "LONG",
            "setup_type": "VWAP reclaim long",
            "signal_quality_score": 76,
            "grade": "A-",
            "action": "TAKE",
            "expected_r_estimate": 0.4,
            "suppression_reasons": [],
            "warnings": [],
        }
    ]
    validation = {
        "report_id": "replay-aware-validation-export-test",
        "model_version": "replay-aware-export-test",
        "validation_mode": "replay_aware_walk_forward",
        "summary": {"selected_candidate_count": 5, "average_r": 0.4},
        "windows": [],
        "selected_trades": [],
        "suppressed_candidates": [],
        "per_symbol": {},
        "per_setup": {},
        "per_regime": {},
        "per_time_bucket": {},
        "sensitivity_summary": {},
        "rejection_reasons": [],
    }

    model_summary = service.export_replay_aware_model_summary_xlsx(model, cells)
    evidence_csv = service.export_evidence_cells_csv("replay-aware-export-test", cells)
    evidence_xlsx = service.export_evidence_cells_xlsx("replay-aware-export-test", cells)
    audits_csv = service.export_score_audits_csv("replay-aware-export-test", audits)
    audits_xlsx = service.export_score_audits_xlsx("replay-aware-export-test", audits)
    validation_xlsx = service.export_replay_aware_validation_xlsx(validation)

    assert {
        "Summary",
        "Training Replay Runs",
        "Evidence Overview",
        "Top Positive Evidence Cells",
        "Top Negative Evidence Cells",
        "Suppression Rules",
        "Scoring Config",
        "Activation Criteria",
        "Warnings",
        "Provenance",
    } <= _sheet_names(model_summary)
    assert evidence_csv.exists()
    assert {
        "Evidence Cells",
        "By Symbol",
        "By Setup",
        "By Regime",
        "By Time Bucket",
        "Fragility Flags",
        "Divergence Flags",
    } <= _sheet_names(evidence_xlsx)
    assert audits_csv.exists()
    assert {"Score Audits"} <= _sheet_names(audits_xlsx)
    assert {
        "Summary",
        "Walk Forward Windows",
        "Selected Trades",
        "Suppressed Candidates",
        "Per Symbol",
        "Per Setup",
        "Per Regime",
        "Per Time Bucket",
        "Sensitivity",
        "Drawdown",
        "Rejection Reasons",
        "Config",
    } <= _sheet_names(validation_xlsx)
