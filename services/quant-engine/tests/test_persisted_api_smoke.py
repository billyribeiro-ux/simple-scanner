from __future__ import annotations

import os
import time
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

from app.config import get_settings
from app.data.symbols import normalize_symbol, normalize_symbols
from app.db.repositories import (
    EXPECTED_TABLES,
    _sync_postgres_url,
    get_repository_registry,
    reset_repository_registry,
)
from app.schemas.market import Bar, Quote
from app.utils.time import UTC

ET = ZoneInfo("America/New_York")
SYNTHETIC_START_ET = datetime(2026, 6, 1, 9, 30, tzinfo=ET)
TEST_FMP_SENTINEL = "test-only-fmp-key"
DEFAULT_POSTGRES_AUTH = ":".join(("amd", "amd"))
DEFAULT_TEST_POSTGRES_DB = "adaptive_market_decoder_test"
DEFAULT_POSTGRES_URL = f"postgresql+psycopg://{DEFAULT_POSTGRES_AUTH}@localhost:15432/{DEFAULT_TEST_POSTGRES_DB}"


def _default_postgres_url() -> str:
    return os.environ.get("TEST_DATABASE_URL") or DEFAULT_POSTGRES_URL


def _synthetic_bars(symbol: str, interval: str, count: int = 120) -> list[Bar]:
    minutes_per_bar = {"1min": 1, "5min": 5, "15min": 15}.get(interval, 1)
    seed = (sum(ord(char) for char in symbol) % 11) * 0.15
    bars = []
    for index in range(count):
        timestamp_et = SYNTHETIC_START_ET + timedelta(minutes=index * minutes_per_bar)
        close = 100.0 + seed + index * 0.12
        bars.append(
            Bar(
                symbol=symbol,
                interval=interval,
                timestamp_utc=timestamp_et.astimezone(UTC),
                timestamp_et=timestamp_et,
                open=close - 0.04,
                high=close + 0.22,
                low=close - 0.18,
                close=close,
                volume=2_000 + index * 35,
                source="mock-fmp",
            )
        )
    return bars


class FakeFMPProvider:
    def __init__(self, *_args, **_kwargs) -> None:
        pass

    def capability_matrix(self) -> list[dict[str, object]]:
        return [{"name": "mock-fmp", "transport": "REST", "v1": True}]

    async def health_check(self) -> dict[str, object]:
        return {"status": "ok", "provider": "mock-fmp"}

    async def get_quote(self, symbol: str) -> Quote:
        return (await self.get_batch_quotes([symbol]))[0]

    async def get_batch_quotes(self, symbols: list[str]) -> list[Quote]:
        quotes = []
        quote_time = (SYNTHETIC_START_ET + timedelta(minutes=121)).astimezone(UTC)
        for index, symbol in enumerate(normalize_symbols(symbols)):
            quotes.append(
                Quote(
                    symbol=symbol,
                    price=114.75 + index,
                    timestamp_utc=quote_time,
                    volume=12_000 + index * 100,
                    source="mock-fmp",
                    raw={"symbol": symbol, "mock": True},
                )
            )
        return quotes

    async def get_historical_bars(
        self,
        symbol: str,
        interval: str,
        _start: datetime,
        _end: datetime,
    ) -> list[Bar]:
        return _synthetic_bars(normalize_symbol(symbol), interval)


def _reset_scanner(routes_module) -> None:
    scanner = routes_module.scanner_state
    scanner.running = False
    scanner.started_at = None
    scanner.last_error = None
    scanner.latest_signals = []
    scanner.context_bars = {}
    scanner.minimum_context_bars = 5
    scanner.scanner_run_id = None
    scanner._task = None
    scanner._queue = None


def _assert_path_clean(path: str | Path) -> None:
    payload = Path(path).read_bytes()
    assert TEST_FMP_SENTINEL.encode() not in payload


def _sheet_names(path: str | Path) -> set[str]:
    workbook = load_workbook(path, read_only=True)
    try:
        return set(workbook.sheetnames)
    finally:
        workbook.close()


def _postgres_available(database_url: str) -> bool:
    try:
        engine = create_engine(_sync_postgres_url(database_url), future=True)
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _clear_postgres_tables(repo) -> None:
    tables = sorted(EXPECTED_TABLES - {"alembic_version"})
    quoted = ", ".join(f'"{table}"' for table in tables)
    with repo.store.engine.begin() as connection:
        connection.execute(text(f"TRUNCATE TABLE {quoted} RESTART IDENTITY CASCADE"))


def _run_persisted_api_vertical_slice(
    tmp_path,
    monkeypatch,
    backend: str,
    database_url: str | None = None,
) -> None:
    db_path = tmp_path / "api-smoke.sqlite3"
    exports_dir = tmp_path / "exports"
    model_dir = tmp_path / "models"
    monkeypatch.setenv("AMD_SQLITE_PATH", str(db_path))
    if backend == "postgresql":
        monkeypatch.setenv("DATABASE_URL", database_url or DEFAULT_POSTGRES_URL)
        monkeypatch.setenv("TEST_DATABASE_URL", database_url or DEFAULT_POSTGRES_URL)
        monkeypatch.setenv("AMD_DB_ROLE", "test")
        monkeypatch.delenv("AMD_ALLOW_SQLITE_FALLBACK", raising=False)
    else:
        monkeypatch.delenv("DATABASE_URL", raising=False)
        monkeypatch.delenv("AMD_DB_ROLE", raising=False)
        monkeypatch.delenv("AMD_ALLOW_SQLITE_FALLBACK", raising=False)
    monkeypatch.setenv("FMP_API_KEY", TEST_FMP_SENTINEL)
    monkeypatch.setenv("PUBLIC_DEFAULT_SYMBOLS", "AAPL,SPY,QQQ,NVDA")

    get_settings.cache_clear()
    reset_repository_registry()
    settings = get_settings()
    monkeypatch.setattr(settings, "exports_dir", exports_dir)
    monkeypatch.setattr(settings, "model_artifacts_dir", model_dir)
    monkeypatch.setattr(settings, "rest_poll_seconds", 60.0)

    from app.api import routes as routes_module
    from app.jobs import scanner as scanner_module
    from app.main import app

    monkeypatch.setattr(routes_module, "FMPMarketDataProvider", FakeFMPProvider)
    monkeypatch.setattr(scanner_module, "FMPMarketDataProvider", FakeFMPProvider)
    _reset_scanner(routes_module)
    repo = get_repository_registry()
    if backend == "postgresql":
        assert repo.info()["persistence_backend"] == "postgresql"
        _clear_postgres_tables(repo)
        assert not db_path.exists()
    else:
        assert repo.info()["persistence_backend"] == "sqlite"

    start = SYNTHETIC_START_ET.astimezone(UTC)
    end = (SYNTHETIC_START_ET + timedelta(minutes=119)).astimezone(UTC)

    with TestClient(app) as client:
        health = client.get("/health").json()
        assert health["persistence"]["persistence_backend"] == backend
        assert health["persistence"]["database_reachable"] is True
        if backend == "sqlite":
            assert health["persistence"]["path"] == str(db_path)
        else:
            assert health["persistence"]["runtime_mode"] == "postgresql"
            assert health["persistence"]["db_role"] == "test"

        config = client.get("/config").json()
        assert config["fmp_api_key_configured"] is True
        assert config["persistence"]["database_url_kind"] == ("postgresql" if backend == "postgresql" else "unset")

        ingest = client.post(
            "/data/ingest",
            json={
                "symbols": ["APPL", "SPY", "QQQ", "NVDA"],
                "intervals": ["1min"],
                "start": start.isoformat(),
                "end": end.isoformat(),
            },
        ).json()
        assert ingest["status"] == "ok"
        assert "AAPL" in ingest["symbols"]
        assert "APPL" not in ingest["symbols"]
        assert ingest["bars_written"] == 480
        assert {
            "features",
            "candidates",
            "labels",
            "replay",
        } <= {row["artifact_type"] for row in ingest["dirty_ranges"]}

        bars = client.get("/data/bars").json()
        assert len(bars) == 480
        quality_report = client.get(
            "/data/quality-report",
            params={"symbols": "AAPL,SPY", "intervals": "1min", "start": start.isoformat(), "end": end.isoformat()},
        ).json()
        assert quality_report["status"] == "ok"
        assert quality_report["summary"]["bar_count"] == 240

        latest_quotes = client.get("/data/quotes/latest").json()
        assert {quote["symbol"] for quote in latest_quotes} == {"AAPL", "SPY", "QQQ", "NVDA"}

        features = client.post("/features/build").json()
        assert features["features"] == 480
        assert features["features_written"] == 480
        assert features["build_windows"]

        labels = client.post("/labels/build").json()
        assert labels["labels"] > 0
        assert labels["candidates"] > 0
        assert labels["build_windows"]

        train = client.post(
            "/models/train",
            json={
                "symbols": ["AAPL", "SPY", "QQQ", "NVDA"],
                "training_start": start.isoformat(),
                "training_end": end.isoformat(),
                "min_samples": 1,
                "activate_if_passes": False,
            },
        ).json()
        model_version = train["model_version"]

        activation_without_report = client.post(
            "/models/activate",
            params={"model_version": model_version},
        ).json()
        assert activation_without_report["activated"] is False
        assert activation_without_report["reason"] == "accepted_validation_report_required"

        validation = client.post("/models/validate", params={"model_version": model_version}).json()
        assert validation["model_version"] == model_version

        repo = get_repository_registry()
        repo.validation_reports.save(
            {
                "model_version": model_version,
                "summary": {"total_trades": 0},
                "windows": [],
                "activation_decision": "rejected",
                "rejection_reasons": ["controlled_rejection"],
                "created_at": (datetime.now(UTC) - timedelta(seconds=2)).isoformat(),
            },
            model_version=model_version,
        )
        activation_with_rejected_report = client.post(
            "/models/activate",
            params={"model_version": model_version},
        ).json()
        assert activation_with_rejected_report["activated"] is False
        assert activation_with_rejected_report["reason"] == "validation_gate_failed"

        repo.validation_reports.save(
            {
                "model_version": model_version,
                "summary": {"total_trades": 40, "average_r": 0.25, "profit_factor": 1.6},
                "windows": [],
                "activation_decision": "accepted",
                "rejection_reasons": [],
                "created_at": datetime.now(UTC).isoformat(),
            },
            model_version=model_version,
        )
        activated = client.post("/models/activate", params={"model_version": model_version}).json()
        assert activated["activated"] is True
        assert activated["active_model"]["model_version"] == model_version

        replacement_version = f"{model_version}-replacement"
        replacement_model = dict(train)
        replacement_model["model_version"] = replacement_version
        replacement_model["active"] = False
        replacement_model["metrics"] = {
            "average_r": -0.1,
            "profit_factor": 0.5,
            "max_drawdown_r": -3.0,
            "observed_outcome_count": 5,
            "total_trades": 5,
        }
        replacement_model["validation_metrics"] = {}
        repo.model_runs.save(replacement_model, artifact_path=str(model_dir / f"{replacement_version}.json"))
        repo.validation_reports.save(
            {
                "model_version": replacement_version,
                "summary": {"total_trades": 40, "average_r": 0.25, "profit_factor": 1.6},
                "windows": [],
                "activation_decision": "accepted",
                "rejection_reasons": [],
                "created_at": (datetime.now(UTC) + timedelta(seconds=1)).isoformat(),
            },
            model_version=replacement_version,
        )
        replacement = client.post("/models/activate", params={"model_version": replacement_version}).json()
        assert replacement["activated"] is True
        assert replacement["active_model"]["model_version"] == replacement_version

        models = client.get("/models").json()
        assert {model["model_version"] for model in models} >= {model_version, replacement_version}
        model_detail = client.get(f"/models/{replacement_version}").json()
        assert model_detail["model_version"] == replacement_version

        active_model_runs = [model for model in repo.model_runs.list_all() if model.get("active")]
        assert len(active_model_runs) == 1

        backtest = client.post(
            "/backtest/run",
            json={
                "symbols": ["AAPL", "SPY", "QQQ", "NVDA"],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "model_version": model_version,
            },
        ).json()
        assert backtest["report_id"]
        assert backtest["simulation_type"] == "label_derived"
        assert backtest["summary"]["total_trades"] > 0
        backtest_runs = client.get("/backtest/runs").json()
        assert backtest["report_id"] in {run["report_id"] for run in backtest_runs}
        backtest_run = client.get(f"/backtest/runs/{backtest['report_id']}").json()
        assert backtest_run["report_id"] == backtest["report_id"]
        assert backtest_run["simulation_type"] == "label_derived"

        replay = client.post(
            "/backtest/replay",
            json={
                "symbols": ["AAPL", "SPY", "QQQ", "NVDA"],
                "intervals": ["1min"],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "max_hold_minutes": 30,
                "minimum_reward_risk": 0.5,
                "feature_warmup_bars": 1,
            },
        ).json()
        replay_run_id = replay["replay_run_id"]
        assert replay["simulation_type"] == "candidate_market_replay"
        assert replay["config_hash"]
        assert replay["input_fingerprint"]
        assert replay["candidate_fingerprint"]
        assert replay["summary_metrics"]["total_candidates"] > 0
        assert replay["trades_written"] == replay["summary_metrics"]["total_candidates"]
        pipeline_status = client.get("/pipeline/status").json()
        assert "dirty_window_count" in pipeline_status
        replay_detail = client.get(f"/backtest/replay/{replay_run_id}").json()
        assert replay_detail["replay_run_id"] == replay_run_id
        replay_trades = client.get(
            f"/backtest/replay/{replay_run_id}/trades",
            params={"limit": 100_000},
        ).json()
        assert replay_trades["trades"]
        assert {trade["status"] for trade in replay_trades["trades"]} <= {"TAKEN", "SKIPPED"}
        replay_trade_count = len(replay_trades["trades"])
        taken_replay_trades = client.get(
            f"/backtest/replay/{replay_run_id}/trades",
            params={"status": "TAKEN"},
        ).json()
        assert all(trade["status"] == "TAKEN" for trade in taken_replay_trades["trades"])
        counterfactual_replay = client.post(
            "/backtest/replay",
            json={
                "symbols": ["AAPL", "SPY", "QQQ", "NVDA"],
                "intervals": ["1min"],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "session": "rth",
                "max_hold_minutes": 15,
                "replay_purpose": "model_training_counterfactual",
                "minimum_reward_risk": 0.5,
                "allow_stale": False,
                "feature_warmup_bars": 1,
            },
        ).json()
        counterfactual_replay_run_id = counterfactual_replay["replay_run_id"]
        assert counterfactual_replay["simulation_type"] == "model_training_counterfactual"
        assert counterfactual_replay["summary_metrics"]["candidate_quality_mode"] is True
        counterfactual_trades = client.get(
            f"/backtest/replay/{counterfactual_replay_run_id}/trades",
            params={"limit": 100_000},
        ).json()
        assert counterfactual_trades["trades"]
        assert any((trade.get("metadata") or {}).get("counterfactual_observed") for trade in counterfactual_trades["trades"])
        combined_runs = client.get("/backtest/runs").json()
        assert replay_run_id in {
            run.get("replay_run_id") or run.get("report_id")
            for run in combined_runs
        }
        replay_run = client.get(f"/backtest/runs/{replay_run_id}").json()
        assert replay_run["replay_run_id"] == replay_run_id
        missing_replay_selection = client.post(
            "/models/validate",
            params={"model_version": model_version, "validation_mode": "candidate_market_replay"},
        ).json()
        assert "replay_run_selection_required" in missing_replay_selection["rejection_reasons"]
        replay_validation = client.post(
            "/models/validate",
            params={
                "model_version": model_version,
                "validation_mode": "candidate_market_replay",
                "replay_run_id": replay_run_id,
            },
        ).json()
        assert replay_validation["replay_run_id"] == replay_run_id
        assert replay_validation["replay_run_selection"]["reason"] == "explicit_replay_run_id"
        sensitivity = client.post(
            f"/backtest/replay/{replay_run_id}/sensitivity",
            json={
                "slippage_bps_grid": [0, 1],
                "spread_bps_grid": [0, 2],
                "intrabar_path_policies": ["conservative"],
                "minimum_total_trades": 1,
                "minimum_robustness_score": 0,
            },
        ).json()
        sensitivity_run_id = sensitivity["sensitivity_run_id"]
        assert sensitivity["status"] == "ok"
        assert sensitivity["scenario_count"] == 4
        sensitivity_detail = client.get(f"/backtest/replay/sensitivity/{sensitivity_run_id}").json()
        assert sensitivity_detail["sensitivity_run_id"] == sensitivity_run_id
        sensitivity_scenarios = client.get(
            f"/backtest/replay/sensitivity/{sensitivity_run_id}/scenarios"
        ).json()
        assert len(sensitivity_scenarios["scenarios"]) == 4
        sensitivity_list = client.get(f"/backtest/replay/{replay_run_id}/sensitivity").json()
        assert sensitivity_run_id in {
            run["sensitivity_run_id"] for run in sensitivity_list["sensitivity_runs"]
        }
        comparison = client.post(
            "/backtest/compare-label-vs-replay",
            json={"replay_run_id": replay_run_id},
        ).json()
        assert comparison["status"] == "ok"
        comparison_detail = client.get(f"/backtest/comparisons/{comparison['comparison_id']}").json()
        assert comparison_detail["comparison_id"] == comparison["comparison_id"]
        counterfactual_comparison = client.post(
            "/backtest/compare-counterfactual-vs-portfolio",
            json={
                "counterfactual_replay_run_id": counterfactual_replay_run_id,
                "portfolio_replay_run_id": replay_run_id,
            },
        ).json()
        assert counterfactual_comparison["status"] == "ok"
        counterfactual_comparison_detail = client.get(
            f"/backtest/counterfactual-comparisons/{counterfactual_comparison['comparison_id']}"
        ).json()
        assert counterfactual_comparison_detail["comparison_type"] == "counterfactual_vs_portfolio"

        replay_aware_train = client.post(
            "/models/train",
            json={
                "model_type": "replay_aware_baseline",
                "symbols": ["AAPL", "SPY", "QQQ", "NVDA"],
                "intervals": ["1min"],
                "training_start": start.isoformat(),
                "training_end": end.isoformat(),
                "min_samples": 1,
                "counterfactual_replay_run_ids": [counterfactual_replay_run_id],
                "portfolio_replay_run_ids": [replay_run_id],
                "outcome_source": "counterfactual_preferred",
                "require_counterfactual": True,
                "minimum_observed_outcomes": 1,
                "minimum_cell_sample_size": 1,
                "scoring_config": {"take_score_threshold": 50},
                "activation_criteria": {
                    "minimum_selected_trades": 1,
                    "minimum_profit_factor": 0.0,
                    "maximum_symbol_profit_share": 1.0,
                    "maximum_setup_profit_share": 1.0,
                },
            },
        ).json()
        assert replay_aware_train["trained"] is True
        replay_aware_model_version = replay_aware_train["model_version"]
        evidence = client.get(f"/models/{replay_aware_model_version}/evidence").json()
        assert evidence["summary"]["cell_count"] > 0
        inline_candidate = dict(replay_trades["trades"][0])
        inline_candidate["timestamp_utc"] = inline_candidate["signal_timestamp_utc"]
        scored_candidates = client.post(
            f"/models/{replay_aware_model_version}/score-candidates",
            json={"candidates": [inline_candidate], "persist_audit": True},
        ).json()
        assert scored_candidates["status"] == "ok"
        assert scored_candidates["scores"]
        score_audits = client.get(f"/models/{replay_aware_model_version}/score-audits").json()
        assert score_audits["score_audits"]
        calibration_audit = client.post(
            f"/models/{replay_aware_model_version}/calibration-audit",
            json={
                "replay_run_ids": [counterfactual_replay_run_id],
                "outcome_source": "counterfactual_only",
                "minimum_high_grade_samples": 1,
            },
        ).json()
        calibration_audit_id = calibration_audit["calibration_audit_id"]
        assert calibration_audit["status"] == "ok"
        assert calibration_audit["model_version"] == replay_aware_model_version
        calibration_list = client.get(f"/models/{replay_aware_model_version}/calibration-audits").json()
        assert calibration_audit_id in {audit["calibration_audit_id"] for audit in calibration_list["calibration_audits"]}
        calibration_detail = client.get(f"/models/calibration-audits/{calibration_audit_id}").json()
        assert calibration_detail["calibration_audit_id"] == calibration_audit_id
        calibration_bins = client.get(f"/models/calibration-audits/{calibration_audit_id}/bins").json()
        assert calibration_bins["bins"]
        model_comparison = client.post(
            "/models/compare",
            json={
                "model_versions": [replay_aware_model_version],
                "calibration_audit_ids": [calibration_audit_id],
                "replay_run_ids": [counterfactual_replay_run_id, replay_run_id],
            },
        ).json()
        assert model_comparison["status"] == "ok"
        replay_aware_validation = client.post(
            "/models/validate",
            params={
                "model_version": replay_aware_model_version,
                "validation_mode": "replay_aware_walk_forward",
                "training_replay_run_ids": counterfactual_replay_run_id,
                "validation_replay_run_ids": counterfactual_replay_run_id,
                "calibration_audit_id": calibration_audit_id,
            },
        ).json()
        assert replay_aware_validation["validation_mode"] == "replay_aware_walk_forward"
        window_set = client.post(
            "/orchestration/replay-window-sets",
            json={
                "name": "api-smoke-window-set",
                "symbols": ["AAPL", "SPY"],
                "intervals": ["1min"],
                "window_mode": "custom",
                "windows": [
                    {
                        "replay_start": start.isoformat(),
                        "replay_end": (start + timedelta(minutes=30)).isoformat(),
                    }
                ],
                "replay_config": {"allow_stale": True, "max_hold_minutes": 15},
                "model_version": replay_aware_model_version,
                "run_immediately": False,
            },
        ).json()
        window_set_id = window_set["window_set_id"]
        assert window_set["status"] == "created"
        assert window_set["symbols"] == ["AAPL", "SPY"]
        assert client.get("/orchestration/replay-window-sets").json()["window_sets"]
        assert client.get(f"/orchestration/replay-window-sets/{window_set_id}").json()["window_set_id"] == window_set_id
        window_run = client.post(
            f"/orchestration/replay-window-sets/{window_set_id}/run",
            json={"run_replay": False, "run_calibration": False},
        ).json()
        assert window_run["status"] == "ok"
        assert window_run["summary"]["completed_window_count"] == 1
        window_results = client.get(f"/orchestration/replay-window-sets/{window_set_id}/results").json()
        assert len(window_results["results"]) == 1
        window_set_export_direct = client.post(f"/orchestration/replay-window-sets/{window_set_id}/export").json()
        assert window_set_export_direct["status"] == "ok"
        calibration_drift = client.post(
            f"/models/{replay_aware_model_version}/calibration-drift",
            json={
                "calibration_audit_ids": [calibration_audit_id],
                "window_set_id": window_set_id,
                "minimum_recent_high_grade_samples": 1,
            },
        ).json()
        drift_report_id = calibration_drift["drift_report_id"]
        assert calibration_drift["status"] == "ok"
        drift_list = client.get(f"/models/{replay_aware_model_version}/calibration-drift").json()
        assert drift_report_id in {report["drift_report_id"] for report in drift_list["drift_reports"]}
        drift_detail = client.get(f"/models/calibration-drift/{drift_report_id}").json()
        assert drift_detail["drift_report_id"] == drift_report_id
        drift_windows = client.get(f"/models/calibration-drift/{drift_report_id}/windows").json()
        assert drift_windows["windows"]
        model_review = client.post(
            f"/models/{replay_aware_model_version}/review-report",
            json={
                "validation_report_ids": [replay_aware_validation["report_id"]],
                "calibration_audit_ids": [calibration_audit_id],
                "drift_report_ids": [drift_report_id],
                "window_set_id": window_set_id,
            },
        ).json()
        review_report_id = model_review["review_report_id"]
        assert model_review["status"] == "ok"
        assert model_review["summary"]["model_activation_unchanged"] is True
        review_list = client.get(f"/models/{replay_aware_model_version}/review-reports").json()
        assert review_report_id in {report["review_report_id"] for report in review_list["review_reports"]}
        assert client.get(f"/models/review-reports/{review_report_id}").json()["review_report_id"] == review_report_id
        accepted_replay_validation = repo.validation_reports.save(
            {
                "model_version": replay_aware_model_version,
                "validation_mode": "replay_aware_walk_forward",
                "summary": {"selected_candidate_count": 5, "average_r": 0.25, "profit_factor": 1.5},
                "windows": [],
                "activation_decision": "accepted",
                "rejection_reasons": [],
                "created_at": datetime.now(UTC).isoformat(),
            },
            model_version=replay_aware_model_version,
            purpose="replay_aware_validation",
        )
        refreshed_model_review = client.post(
            f"/models/{replay_aware_model_version}/review-report",
            json={
                "validation_report_ids": [accepted_replay_validation["report_id"]],
                "calibration_audit_ids": [calibration_audit_id],
                "drift_report_ids": [drift_report_id],
                "window_set_id": window_set_id,
            },
        ).json()
        assert refreshed_model_review["status"] == "ok"
        review_report_id = refreshed_model_review["review_report_id"]
        research_cycle = client.post(
            "/research/cycles",
            json={
                "cycle_date": "2026-07-01",
                "cycle_type": "daily",
                "symbols": ["AAPL", "SPY"],
                "intervals": ["1min"],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "active_model_version": replacement_version,
                "challenger_model_version": replay_aware_model_version,
                "counterfactual_replay_run_ids": [counterfactual_replay_run_id],
                "portfolio_replay_run_ids": [replay_run_id],
                "sensitivity_run_ids": [sensitivity_run_id],
                "validation_report_ids": [accepted_replay_validation["report_id"]],
                "calibration_audit_ids": [calibration_audit_id],
                "drift_report_ids": [drift_report_id],
                "model_review_report_ids": [review_report_id],
                "allow_stale": True,
            },
        ).json()
        research_cycle_id = research_cycle["research_cycle_id"]
        assert research_cycle["status"] == "created"
        cycle_dry_run = client.post(f"/research/cycles/{research_cycle_id}/dry-run").json()
        assert cycle_dry_run["status"] == "dry_run"
        cycle_run = client.post(f"/research/cycles/{research_cycle_id}/run", json={"allow_stale": True}).json()
        assert cycle_run["status"] in {"completed", "blocked"}
        assert cycle_run["summary"]["model_activation_unchanged"] is True
        cycle_artifacts = client.get(f"/research/cycles/{research_cycle_id}/artifacts").json()
        assert cycle_artifacts["artifacts"]
        cycles = client.get("/research/cycles").json()
        assert research_cycle_id in {cycle["research_cycle_id"] for cycle in cycles["research_cycles"]}
        cycle_detail = client.get(f"/research/cycles/{research_cycle_id}").json()
        assert cycle_detail["research_cycle_id"] == research_cycle_id
        proposal_id = cycle_run["proposal"]["proposal_id"]
        assert cycle_run["proposal"]["recommended_action"] == "APPROVE_CHALLENGER_FOR_ACTIVATION"
        proposals = client.get("/research/model-proposals").json()
        assert proposal_id in {proposal["proposal_id"] for proposal in proposals["model_proposals"]}
        proposal_detail = client.get(f"/research/model-proposals/{proposal_id}").json()
        assert proposal_detail["proposal_id"] == proposal_id
        approved_proposal = client.post(
            f"/research/model-proposals/{proposal_id}/approve",
            json={"actor": "api-smoke"},
        ).json()
        assert approved_proposal["status"] == "APPROVED_FOR_ACTIVATION"
        proposal_activation_without_confirmation = client.post(
            f"/research/model-proposals/{proposal_id}/activate",
            json={"actor": "api-smoke", "confirm_manual_activation": False},
        ).json()
        assert proposal_activation_without_confirmation["status"] == "blocked"
        decision_ledger = client.get("/research/decision-ledger", params={"proposal_id": proposal_id}).json()
        assert decision_ledger["decisions"]
        research_status = client.get("/operations/research-status").json()
        assert research_status["status"] == "ok"
        assert research_status["latest_research_cycle"]["research_cycle_id"] == research_cycle_id
        scheduler_job = client.post(
            "/scheduler/jobs",
            json={
                "job_type": "data_quality_report",
                "payload": {"symbols": ["AAPL", "SPY"], "intervals": ["1min"]},
                "created_by": "api-smoke",
            },
        ).json()
        scheduler_export_job = client.post(
            "/scheduler/jobs",
            json={"job_type": "export_operations_status", "payload": {}, "created_by": "api-smoke"},
        ).json()
        scheduler_cancel_job = client.post(
            "/scheduler/jobs",
            json={"job_type": "data_quality_report", "payload": {"symbols": ["QQQ"]}},
        ).json()
        assert scheduler_job["status"] == "QUEUED"
        assert scheduler_export_job["status"] == "QUEUED"
        assert client.post(f"/scheduler/jobs/{scheduler_cancel_job['job_id']}/cancel").json()["status"] == "CANCELLED"
        scheduler_list = client.get("/scheduler/jobs").json()
        assert scheduler_job["job_id"] in {job["job_id"] for job in scheduler_list["jobs"]}
        scheduler_pending = client.post("/scheduler/jobs/run-pending", json={"max_jobs": 2}).json()
        assert scheduler_pending["jobs_run"] == 2
        assert {result["status"] for result in scheduler_pending["results"]} == {"COMPLETED"}
        scheduler_export_result = next(
            result for result in scheduler_pending["results"] if result["job_id"] == scheduler_export_job["job_id"]
        )
        scheduler_export_path = scheduler_export_result["result"]["path"]
        scheduler_events = client.get(f"/scheduler/jobs/{scheduler_job['job_id']}/events").json()
        assert scheduler_events["events"]
        scheduler_status = client.get("/operations/scheduler-status").json()
        assert scheduler_status["completed_jobs"] >= 2
        assert scheduler_status["cancelled_jobs"] >= 1
        research_status_with_scheduler = client.get("/operations/research-status").json()
        assert research_status_with_scheduler["queued_scheduler_jobs"] == 0
        repo.validation_reports.save(
            {
                "model_version": replay_aware_model_version,
                "validation_mode": "replay_aware_walk_forward",
                "summary": {"selected_candidate_count": 5, "average_r": 0.25, "profit_factor": 1.5},
                "windows": [],
                "activation_decision": "accepted",
                "rejection_reasons": [],
                "created_at": datetime.now(UTC).isoformat(),
            },
            model_version=replay_aware_model_version,
            purpose="replay_aware_validation",
        )
        proposal_activation = client.post(
            f"/research/model-proposals/{proposal_id}/activate",
            json={
                "actor": "api-smoke",
                "confirm_manual_activation": True,
                "validation_mode": "replay_aware_walk_forward",
            },
        ).json()
        assert proposal_activation["status"] == "ok"
        assert proposal_activation["activation"]["activated"] is True
        replay_aware_activation = client.post(
            "/models/activate",
            params={
                "model_version": replay_aware_model_version,
                "validation_mode": "replay_aware_walk_forward",
            },
        ).json()
        assert replay_aware_activation["activated"] is True

        scanner_start = client.post(
            "/scanner/start",
            json={"symbols": ["AAPL", "SPY"], "confidence_threshold": 0.0},
        ).json()
        assert scanner_start["scanner_run_id"]
        live_signals = []
        for _ in range(40):
            live_signals = client.get("/signals/live").json()
            if live_signals:
                break
            time.sleep(0.05)
        scanner_status = client.get("/scanner/status").json()
        scanner_stop = client.post("/scanner/stop").json()
        assert scanner_stop["running"] is False
        assert scanner_status["latest_persisted_run"]["scanner_run_id"] == scanner_start["scanner_run_id"]
        assert live_signals
        signal_history = client.get("/signals/history").json()
        assert len(signal_history) == len(live_signals)

        csv_export = client.post("/exports/signals.csv", json={"kind": "signals"}).json()
        xlsx_export = client.post("/exports/signals.xlsx", json={"kind": "signals"}).json()
        backtest_export = client.post("/exports/backtest.xlsx", json={"kind": "backtest"}).json()
        replay_summary_export = client.post(
            "/exports/replay-summary.xlsx",
            json={"kind": "replay-summary", "run_id": replay_run_id},
        ).json()
        replay_trades_csv = client.post(
            "/exports/replay-trades.csv",
            json={"kind": "replay-trades", "run_id": replay_run_id},
        ).json()
        replay_trades_xlsx = client.post(
            "/exports/replay-trades.xlsx",
            json={"kind": "replay-trades", "run_id": replay_run_id},
        ).json()
        sensitivity_summary_export = client.post(
            "/exports/sensitivity-summary.xlsx",
            json={"kind": "sensitivity-summary", "run_id": sensitivity_run_id},
        ).json()
        sensitivity_scenarios_csv = client.post(
            "/exports/sensitivity-scenarios.csv",
            json={"kind": "sensitivity-scenarios", "run_id": sensitivity_run_id},
        ).json()
        sensitivity_scenarios_xlsx = client.post(
            "/exports/sensitivity-scenarios.xlsx",
            json={"kind": "sensitivity-scenarios", "run_id": sensitivity_run_id},
        ).json()
        sensitivity_metrics_export = client.post(
            "/exports/sensitivity-metrics.json",
            json={"kind": "sensitivity-metrics", "run_id": sensitivity_run_id},
        ).json()
        replay_aware_model_export = client.post(
            "/exports/replay-aware-model-summary.xlsx",
            json={"kind": "replay-aware-model-summary", "run_id": replay_aware_model_version},
        ).json()
        evidence_cells_csv = client.post(
            "/exports/evidence-cells.csv",
            json={"kind": "evidence-cells", "run_id": replay_aware_model_version},
        ).json()
        evidence_cells_xlsx = client.post(
            "/exports/evidence-cells.xlsx",
            json={"kind": "evidence-cells", "run_id": replay_aware_model_version},
        ).json()
        score_audits_csv = client.post(
            "/exports/score-audits.csv",
            json={"kind": "score-audits", "run_id": replay_aware_model_version},
        ).json()
        score_audits_xlsx = client.post(
            "/exports/score-audits.xlsx",
            json={"kind": "score-audits", "run_id": replay_aware_model_version},
        ).json()
        replay_aware_validation_export = client.post(
            "/exports/replay-aware-validation.xlsx",
            json={"kind": "replay-aware-validation", "run_id": replay_aware_validation["report_id"]},
        ).json()
        calibration_audit_export = client.post(
            "/exports/calibration-audit.xlsx",
            json={"kind": "calibration-audit", "run_id": calibration_audit_id},
        ).json()
        calibration_bins_csv = client.post(
            "/exports/calibration-bins.csv",
            json={"kind": "calibration-bins", "run_id": calibration_audit_id},
        ).json()
        calibration_bins_xlsx = client.post(
            "/exports/calibration-bins.xlsx",
            json={"kind": "calibration-bins", "run_id": calibration_audit_id},
        ).json()
        calibration_metrics_json = client.post(
            "/exports/calibration-metrics.json",
            json={"kind": "calibration-metrics", "run_id": calibration_audit_id},
        ).json()
        model_comparison_export = client.post(
            "/exports/model-comparison.xlsx",
            json={"kind": "model-comparison", "run_id": model_comparison["comparison_id"]},
        ).json()
        replay_window_set_export = client.post(
            "/exports/replay-window-set.xlsx",
            json={"kind": "replay-window-set", "run_id": window_set_id},
        ).json()
        drift_export_xlsx = client.post(
            "/exports/calibration-drift.xlsx",
            json={"kind": "calibration-drift", "run_id": drift_report_id},
        ).json()
        drift_export_json = client.post(
            "/exports/calibration-drift.json",
            json={"kind": "calibration-drift", "run_id": drift_report_id},
        ).json()
        drift_windows_csv = client.post(
            "/exports/calibration-drift-windows.csv",
            json={"kind": "calibration-drift-windows", "run_id": drift_report_id},
        ).json()
        drift_windows_xlsx = client.post(
            "/exports/calibration-drift-windows.xlsx",
            json={"kind": "calibration-drift-windows", "run_id": drift_report_id},
        ).json()
        model_review_xlsx = client.post(
            "/exports/model-review.xlsx",
            json={"kind": "model-review", "run_id": review_report_id},
        ).json()
        model_review_json = client.post(
            "/exports/model-review.json",
            json={"kind": "model-review", "run_id": review_report_id},
        ).json()
        research_cycle_xlsx = client.post(
            "/exports/research-cycle.xlsx",
            json={"kind": "research-cycle", "run_id": research_cycle_id},
        ).json()
        research_cycle_json = client.post(
            "/exports/research-cycle.json",
            json={"kind": "research-cycle", "run_id": research_cycle_id},
        ).json()
        model_proposal_xlsx = client.post(
            "/exports/model-proposal.xlsx",
            json={"kind": "model-proposal", "run_id": proposal_id},
        ).json()
        model_proposal_json = client.post(
            "/exports/model-proposal.json",
            json={"kind": "model-proposal", "run_id": proposal_id},
        ).json()
        champion_challenger_xlsx = client.post(
            "/exports/champion-challenger-comparison.xlsx",
            json={"kind": "champion-challenger-comparison", "run_id": cycle_run["comparison"]["comparison_id"]},
        ).json()
        review = client.post("/review/daily").json()
        persisted_review = client.get(f"/review/daily/{review['date']}").json()
        daily_export = client.post("/exports/daily-review.xlsx", json={"kind": "daily-review"}).json()
        export_status = client.get(f"/exports/{csv_export['export']['export_id']}").json()

        assert csv_export["rows"] == len(live_signals)
        assert xlsx_export["rows"] == len(live_signals)
        assert backtest_export["status"] == "ok"
        assert backtest_export["note"] == "V1 workbook scaffold"
        assert replay_summary_export["status"] == "ok"
        assert len(replay_summary_export["paths"]) == 2
        assert replay_trades_csv["rows"] == replay_trade_count
        assert replay_trades_xlsx["rows"] == replay_trade_count
        assert sensitivity_summary_export["status"] == "ok"
        assert sensitivity_summary_export["export"]["file_sha256"]
        assert sensitivity_scenarios_csv["rows"] == 4
        assert sensitivity_scenarios_xlsx["rows"] == 4
        assert sensitivity_metrics_export["rows"] == 4
        assert replay_aware_model_export["status"] == "ok"
        assert evidence_cells_csv["rows"] == evidence["summary"]["cell_count"]
        assert evidence_cells_xlsx["rows"] == evidence["summary"]["cell_count"]
        assert score_audits_csv["rows"] >= 1
        assert score_audits_xlsx["rows"] >= 1
        assert replay_aware_validation_export["status"] == "ok"
        assert calibration_audit_export["status"] == "ok"
        assert calibration_bins_csv["rows"] == len(calibration_bins["bins"])
        assert calibration_bins_xlsx["rows"] == len(calibration_bins["bins"])
        assert calibration_metrics_json["status"] == "ok"
        assert model_comparison_export["status"] == "ok"
        assert window_set_export_direct["status"] == "ok"
        assert replay_window_set_export["status"] == "ok"
        assert drift_export_xlsx["status"] == "ok"
        assert drift_export_json["status"] == "ok"
        assert drift_windows_csv["rows"] == len(drift_windows["windows"])
        assert drift_windows_xlsx["rows"] == len(drift_windows["windows"])
        assert model_review_xlsx["status"] == "ok"
        assert model_review_json["status"] == "ok"
        assert research_cycle_xlsx["status"] == "ok"
        assert research_cycle_json["status"] == "ok"
        assert model_proposal_xlsx["status"] == "ok"
        assert model_proposal_json["status"] == "ok"
        assert champion_challenger_xlsx["status"] == "ok"
        assert {"Live Signals", "Model Info"} <= _sheet_names(xlsx_export["path"])
        assert {"Live Signals", "Model Info"} <= _sheet_names(backtest_export["path"])
        assert {
            "Summary",
            "Trades",
            "Skipped Candidates",
            "Per Symbol",
            "Per Setup",
            "Config",
            "Warnings",
        } <= _sheet_names(replay_summary_export["paths"][0])
        assert {"Trades"} <= _sheet_names(replay_trades_xlsx["path"])
        assert {
            "Summary",
            "Scenario Metrics",
            "Worst Case",
            "Median Case",
            "Best Case",
            "Fragility Flags",
            "Gate Results",
            "Config",
            "Warnings",
        } <= _sheet_names(sensitivity_summary_export["path"])
        assert {"Scenario Metrics"} <= _sheet_names(sensitivity_scenarios_xlsx["path"])
        assert {"Summary", "Training Replay Runs", "Evidence Overview"} <= _sheet_names(replay_aware_model_export["path"])
        assert {"Evidence Cells", "By Symbol", "By Setup"} <= _sheet_names(evidence_cells_xlsx["path"])
        assert {"Score Audits"} <= _sheet_names(score_audits_xlsx["path"])
        assert {"Summary", "Walk Forward Windows", "Rejection Reasons"} <= _sheet_names(replay_aware_validation_export["path"])
        assert {"Summary", "Score Bins", "Grade Bins", "Action Bins", "Provenance"} <= _sheet_names(calibration_audit_export["path"])
        assert {"Calibration Bins"} <= _sheet_names(calibration_bins_xlsx["path"])
        assert {"Summary", "Models", "Calibration Metrics", "Warnings"} <= _sheet_names(model_comparison_export["path"])
        assert {"Summary", "Generated Windows", "Window Results", "Config"} <= _sheet_names(replay_window_set_export["path"])
        assert {"Summary", "Drift Flags", "Window Metrics", "Stability"} <= _sheet_names(drift_export_xlsx["path"])
        assert {"Drift Windows"} <= _sheet_names(drift_windows_xlsx["path"])
        assert {"Summary", "Readiness", "Readiness Reasons", "Unresolved Warnings"} <= _sheet_names(model_review_xlsx["path"])
        assert {"Summary", "Cycle Config", "Data Quality", "Champion vs Challenger", "Proposal"} <= _sheet_names(research_cycle_xlsx["path"])
        assert {"Summary", "Recommended Action", "Readiness", "Approval History"} <= _sheet_names(model_proposal_xlsx["path"])
        assert {"Summary", "Champion", "Challenger", "Delta Metrics", "Gates"} <= _sheet_names(champion_challenger_xlsx["path"])
        assert review["signals_fired"] + review["signals_skipped"] == len(live_signals)
        assert persisted_review["status"] == "local-file"
        assert len(daily_export["paths"]) == 3
        assert {"Summary", "Signals Fired", "Recommendations"} <= _sheet_names(daily_export["paths"][2])
        assert export_status["status"] == "local-file"

    reset_repository_registry()
    reopened = get_repository_registry()
    if backend == "sqlite":
        assert reopened.info()["path"] == str(db_path)
    else:
        assert reopened.info()["persistence_backend"] == "postgresql"
        assert not db_path.exists()
    assert len(reopened.bars.list_all()) == 480
    assert len(reopened.features.list_all()) == 480
    assert len(reopened.labels.list_all()) > 0
    assert reopened.replays.get(replay_run_id)["simulation_type"] == "candidate_market_replay"
    assert len(reopened.replays.list_trades(replay_run_id, limit=100_000)) == replay_trade_count
    assert reopened.replay_sensitivity.get(sensitivity_run_id)["replay_run_id"] == replay_run_id
    assert reopened.backtest_comparisons.get(comparison["comparison_id"])["replay_run_id"] == replay_run_id
    assert reopened.backtest_comparisons.get(counterfactual_comparison["comparison_id"])["comparison_type"] == "counterfactual_vs_portfolio"
    assert reopened.model_evidence_cells.count(replay_aware_model_version) == evidence["summary"]["cell_count"]
    assert reopened.candidate_score_audits.list(replay_aware_model_version)
    assert reopened.model_calibration_audits.get(calibration_audit_id)["model_version"] == replay_aware_model_version
    assert reopened.model_comparisons.get(model_comparison["comparison_id"])["comparison_type"] == "model_comparison"
    assert reopened.replay_windows.get_set(window_set_id)["status"] == "completed"
    assert reopened.replay_windows.list_results(window_set_id)[0]["status"] == "completed"
    assert reopened.model_calibration_drift.get(drift_report_id)["model_version"] == replay_aware_model_version
    assert reopened.model_calibration_drift.list_windows(drift_report_id)
    assert reopened.model_review_reports.get(review_report_id)["model_version"] == replay_aware_model_version
    assert reopened.research_cycles.get(research_cycle_id)["research_cycle_id"] == research_cycle_id
    assert reopened.research_cycles.list_artifacts(research_cycle_id)
    assert reopened.model_proposals.get(proposal_id)["proposal_id"] == proposal_id
    assert reopened.champion_challenger_comparisons.get(cycle_run["comparison"]["comparison_id"])["comparison_id"] == cycle_run["comparison"]["comparison_id"]
    assert reopened.model_decision_ledger.list(proposal_id=proposal_id)
    assert reopened.scheduler_jobs.get(scheduler_job["job_id"])["status"] == "COMPLETED"
    assert reopened.scheduler_jobs.list_events(scheduler_job["job_id"])
    assert reopened.active_models.get_active()["model_version"] == replacement_version
    assert reopened.scanner_runs.latest()["scanner_run_id"] == scanner_start["scanner_run_id"]
    assert len(reopened.live_signals.list_latest()) == len(live_signals)
    assert len(reopened.exports.list_all()) >= 20
    assert reopened.daily_reviews.get(datetime.now(UTC).date()) is not None
    assert TEST_FMP_SENTINEL not in str(reopened.provider_requests.list_all())

    clean_paths = [
        csv_export["path"],
        xlsx_export["path"],
        backtest_export["path"],
        replay_trades_csv["path"],
        replay_trades_xlsx["path"],
        sensitivity_summary_export["path"],
        sensitivity_scenarios_csv["path"],
        sensitivity_scenarios_xlsx["path"],
        sensitivity_metrics_export["path"],
        replay_aware_model_export["path"],
        evidence_cells_csv["path"],
        evidence_cells_xlsx["path"],
        score_audits_csv["path"],
        score_audits_xlsx["path"],
        replay_aware_validation_export["path"],
        calibration_audit_export["path"],
        calibration_bins_csv["path"],
        calibration_bins_xlsx["path"],
        calibration_metrics_json["path"],
        model_comparison_export["path"],
        window_set_export_direct["path"],
        replay_window_set_export["path"],
        drift_export_xlsx["path"],
        drift_export_json["path"],
        drift_windows_csv["path"],
        drift_windows_xlsx["path"],
        model_review_xlsx["path"],
        model_review_json["path"],
        research_cycle_xlsx["path"],
        research_cycle_json["path"],
        model_proposal_xlsx["path"],
        model_proposal_json["path"],
        champion_challenger_xlsx["path"],
        scheduler_export_path,
        *replay_summary_export["paths"],
        *daily_export["paths"],
        model_dir / "active_model.json",
    ]
    if backend == "sqlite":
        clean_paths.append(db_path)
    for path in clean_paths:
        _assert_path_clean(path)


def test_persisted_api_vertical_slice_sqlite(tmp_path, monkeypatch) -> None:
    _run_persisted_api_vertical_slice(tmp_path, monkeypatch, backend="sqlite")


def test_persisted_api_vertical_slice_postgres(tmp_path, monkeypatch) -> None:
    database_url = _default_postgres_url()
    if not _postgres_available(database_url):
        pytest.skip("local Postgres/TimescaleDB is not available for persisted API smoke")
    _run_persisted_api_vertical_slice(
        tmp_path,
        monkeypatch,
        backend="postgresql",
        database_url=database_url,
    )
