from __future__ import annotations

import csv
import json
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from app.utils.time import UTC

try:
    from openpyxl import Workbook
except ModuleNotFoundError:  # pragma: no cover - compatibility path for no-venv smoke tests
    Workbook = None  # type: ignore[assignment]

from app.config import get_settings
from app.schemas.market import Signal

SIGNAL_COLUMNS = [
    "timestamp",
    "ticker",
    "side",
    "entry_price",
    "stop_price",
    "target_1",
    "target_2",
    "target_3",
    "risk_per_share",
    "reward_risk_to_t1",
    "reward_risk_to_t2",
    "reward_risk_to_t3",
    "expected_r",
    "confidence_score",
    "signal_grade",
    "setup_type",
    "market_regime",
    "ticker_regime",
    "reasons",
    "warnings",
    "historical_sample_size",
    "historical_win_rate",
    "historical_average_r",
    "model_version",
    "training_start",
    "training_end",
    "data_source",
    "status",
    "exit_price",
    "exit_reason",
    "realized_r",
]

REPLAY_TRADE_COLUMNS = [
    "trade_id",
    "replay_run_id",
    "candidate_id",
    "symbol",
    "interval",
    "side",
    "setup_type",
    "signal_timestamp_utc",
    "entry_timestamp_utc",
    "exit_timestamp_utc",
    "entry_price",
    "stop_price",
    "target_1",
    "target_2",
    "target_3",
    "exit_price",
    "exit_reason",
    "realized_r",
    "mfe_r",
    "mae_r",
    "bars_held",
    "minutes_held",
    "same_bar_ambiguous",
    "ambiguity_policy",
    "slippage_bps",
    "spread_bps",
    "commission",
    "market_regime",
    "time_bucket",
    "signal_score",
    "expected_r",
    "status",
    "skip_reason",
    "replay_purpose",
    "candidate_quality_label",
    "counterfactual_observed",
    "portfolio_blocked_in_execution_replay",
    "concurrent_candidate_count_at_signal",
    "overlap_group_id",
    "concurrency_bucket",
]

SENSITIVITY_SCENARIO_COLUMNS = [
    "scenario_id",
    "sensitivity_run_id",
    "replay_run_id",
    "slippage_bps",
    "spread_bps",
    "intrabar_path_policy",
    "same_bar_stop_target_policy",
    "pass_fail",
    "total_trades",
    "average_r",
    "profit_factor",
    "max_drawdown_r",
    "total_r",
    "skip_rate",
]

EVIDENCE_CELL_COLUMNS = [
    "model_version",
    "cell_key",
    "hierarchy_level",
    "parent_cell_key",
    "sample_size",
    "observed_outcome_count",
    "average_r",
    "median_r",
    "profit_factor",
    "max_drawdown_r",
    "robustness_score",
    "evidence_quality_grade",
]

SCORE_AUDIT_COLUMNS = [
    "score_id",
    "model_version",
    "candidate_id",
    "symbol",
    "interval",
    "timestamp_utc",
    "side",
    "setup_type",
    "signal_quality_score",
    "grade",
    "action",
    "expected_r_estimate",
    "suppression_reasons",
    "warnings",
]

CALIBRATION_BIN_COLUMNS = [
    "calibration_audit_id",
    "bin_type",
    "bin_key",
    "sample_size",
    "observed_average_r",
    "observed_win_rate",
    "profit_factor",
    "max_drawdown_r",
    "same_bar_ambiguity_rate",
    "fragility_flag_rate",
]

MODEL_COMPARISON_COLUMNS = [
    "model_version",
    "observed_outcome_count",
    "average_r",
    "profit_factor",
    "active",
]

REPLAY_WINDOW_RESULT_COLUMNS = [
    "window_result_id",
    "window_set_id",
    "window_index",
    "replay_start",
    "replay_end",
    "counterfactual_replay_run_id",
    "portfolio_replay_run_id",
    "status",
    "total_trades",
    "average_r",
    "profit_factor",
    "max_drawdown_r",
    "warning_count",
]

DRIFT_WINDOW_COLUMNS = [
    "drift_report_id",
    "window_result_id",
    "window_index",
    "severity",
    "rank_correlation_score",
    "monotonicity_pass",
    "high_grade_average_r",
    "take_minus_watch_average_r",
    "flags",
]


class ExportService:
    def __init__(self) -> None:
        self.settings = get_settings()
        self.settings.exports_dir.mkdir(parents=True, exist_ok=True)

    def export_signals_csv(self, signals: Iterable[Signal], run_date: date | None = None) -> Path:
        run_date = run_date or datetime.now(UTC).date()
        path = self.settings.exports_dir / f"live_signals_{run_date.isoformat()}.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=SIGNAL_COLUMNS)
            writer.writeheader()
            for signal in signals:
                writer.writerow(self._signal_row(signal))
        return path

    def export_signals_xlsx(self, signals: Iterable[Signal], run_date: date | None = None) -> Path:
        run_date = run_date or datetime.now(UTC).date()
        path = self.settings.exports_dir / f"live_signals_{run_date.isoformat()}.xlsx"
        rows = [self._signal_row(signal) for signal in signals]
        if Workbook is None:
            self._write_minimal_xlsx(
                path,
                {
                    "Live Signals": [SIGNAL_COLUMNS, *[[row.get(column) for column in SIGNAL_COLUMNS] for row in rows]],
                    "Closed Signals": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                    "Performance Summary": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                    "Ticker Breakdown": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                    "Setup Breakdown": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                    "Regime Breakdown": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                    "Model Info": [["status", "message"], ["scaffold", "Data will populate as signals close and backtests run."]],
                },
            )
            return path
        workbook = Workbook()
        live = workbook.active
        live.title = "Live Signals"
        self._write_sheet(live, SIGNAL_COLUMNS, rows)
        for sheet_name in [
            "Closed Signals",
            "Performance Summary",
            "Ticker Breakdown",
            "Setup Breakdown",
            "Regime Breakdown",
            "Model Info",
        ]:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["status", "message"])
            sheet.append(["scaffold", "Data will populate as signals close and backtests run."])
        workbook.save(path)
        return path

    def export_daily_review(self, payload: dict[str, object], run_date: date | None = None) -> tuple[Path, Path, Path]:
        run_date = run_date or datetime.now(UTC).date()
        json_path = self.settings.exports_dir / f"daily_review_{run_date.isoformat()}.json"
        csv_path = self.settings.exports_dir / f"daily_review_{run_date.isoformat()}.csv"
        xlsx_path = self.settings.exports_dir / f"daily_review_{run_date.isoformat()}.xlsx"
        json_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["section", "value"])
            for key, value in payload.items():
                writer.writerow([key, json.dumps(value, default=str)])
        if Workbook is None:
            self._write_minimal_xlsx(
                xlsx_path,
                {
                    sheet_name: [
                        ["section", "value"],
                        [sheet_name, json.dumps(payload.get(sheet_name.lower().replace(" ", "_"), []), default=str)],
                    ]
                    for sheet_name in [
                        "Summary",
                        "Signals Fired",
                        "Missed Moves",
                        "False Positives",
                        "False Negatives",
                        "Ticker Notes",
                        "Regime Notes",
                        "Recommendations",
                    ]
                },
            )
            return json_path, csv_path, xlsx_path
        workbook = Workbook()
        for index, sheet_name in enumerate(
            [
                "Summary",
                "Signals Fired",
                "Missed Moves",
                "False Positives",
                "False Negatives",
                "Ticker Notes",
                "Regime Notes",
                "Recommendations",
            ]
        ):
            sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
            sheet.title = sheet_name
            sheet.append(["section", "value"])
            sheet.append([sheet_name, json.dumps(payload.get(sheet_name.lower().replace(" ", "_"), []), default=str)])
        workbook.save(xlsx_path)
        return json_path, csv_path, xlsx_path

    def export_replay_trades_csv(self, replay_run_id: str, trades: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"replay_trades_{replay_run_id}.csv"
        rows = [self._replay_trade_row(trade) for trade in trades]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=REPLAY_TRADE_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_replay_trades_xlsx(self, replay_run_id: str, trades: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"replay_trades_{replay_run_id}.xlsx"
        rows = [self._replay_trade_row(trade) for trade in trades]
        if Workbook is None:
            self._write_minimal_xlsx(
                path,
                {"Trades": [REPLAY_TRADE_COLUMNS, *[[row.get(column) for column in REPLAY_TRADE_COLUMNS] for row in rows]]},
            )
            return path
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Trades"
        self._write_sheet(sheet, REPLAY_TRADE_COLUMNS, rows)
        workbook.save(path)
        return path

    def export_replay_summary_xlsx(self, replay_run: dict[str, object], trades: Iterable[dict[str, object]]) -> Path:
        replay_run_id = str(replay_run["replay_run_id"])
        path = self.settings.exports_dir / f"replay_summary_{replay_run_id}.xlsx"
        trade_rows = [self._replay_trade_row(trade) for trade in trades]
        skipped_rows = [row for row in trade_rows if row.get("status") == "SKIPPED"]
        metrics = dict(replay_run.get("summary_metrics") or {})
        sheets = self._replay_workbook_sheets(replay_run, metrics, trade_rows, skipped_rows)
        if Workbook is None:
            self._write_minimal_xlsx(path, sheets)
            return path
        workbook = Workbook()
        for index, (sheet_name, rows) in enumerate(sheets.items()):
            sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
            sheet.title = sheet_name
            for row in rows:
                sheet.append([self._cell_value(value) for value in row])
        workbook.save(path)
        return path

    def export_replay_metrics_json(self, replay_run: dict[str, object]) -> Path:
        replay_run_id = str(replay_run["replay_run_id"])
        path = self.settings.exports_dir / f"replay_metrics_{replay_run_id}.json"
        payload = {
            "replay_run_id": replay_run_id,
            "simulation_type": replay_run.get("simulation_type"),
            "created_at": replay_run.get("created_at"),
            "filters": replay_run.get("config") or {},
            "summary_metrics": replay_run.get("summary_metrics") or {},
        }
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        return path

    def export_sensitivity_summary_xlsx(self, sensitivity_run: dict[str, object], scenarios: Iterable[dict[str, object]]) -> Path:
        sensitivity_run_id = str(sensitivity_run["sensitivity_run_id"])
        path = self.settings.exports_dir / f"replay_sensitivity_summary_{sensitivity_run_id}.xlsx"
        scenario_rows = [self._sensitivity_scenario_row(scenario, sensitivity_run_id) for scenario in scenarios]
        sheets = self._sensitivity_workbook_sheets(sensitivity_run, scenario_rows)
        if Workbook is None:
            self._write_minimal_xlsx(path, sheets)
            return path
        workbook = Workbook()
        for index, (sheet_name, rows) in enumerate(sheets.items()):
            sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
            sheet.title = sheet_name
            for row in rows:
                sheet.append([self._cell_value(value) for value in row])
        workbook.save(path)
        return path

    def export_sensitivity_scenarios_csv(self, sensitivity_run_id: str, scenarios: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"replay_sensitivity_scenarios_{sensitivity_run_id}.csv"
        rows = [self._sensitivity_scenario_row(scenario, sensitivity_run_id) for scenario in scenarios]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=SENSITIVITY_SCENARIO_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_sensitivity_scenarios_xlsx(self, sensitivity_run_id: str, scenarios: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"replay_sensitivity_scenarios_{sensitivity_run_id}.xlsx"
        rows = [self._sensitivity_scenario_row(scenario, sensitivity_run_id) for scenario in scenarios]
        if Workbook is None:
            self._write_minimal_xlsx(
                path,
                {"Scenario Metrics": [SENSITIVITY_SCENARIO_COLUMNS, *[[row.get(column) for column in SENSITIVITY_SCENARIO_COLUMNS] for row in rows]]},
            )
            return path
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Metrics"
        self._write_sheet(sheet, SENSITIVITY_SCENARIO_COLUMNS, rows)
        workbook.save(path)
        return path

    def export_sensitivity_metrics_json(self, sensitivity_run: dict[str, object], scenarios: Iterable[dict[str, object]]) -> Path:
        sensitivity_run_id = str(sensitivity_run["sensitivity_run_id"])
        path = self.settings.exports_dir / f"replay_sensitivity_metrics_{sensitivity_run_id}.json"
        payload = {
            "sensitivity_run_id": sensitivity_run_id,
            "replay_run_id": sensitivity_run.get("replay_run_id"),
            "created_at": sensitivity_run.get("created_at"),
            "config": sensitivity_run.get("config") or {},
            "robustness_score": sensitivity_run.get("robustness_score"),
            "pass_fail": sensitivity_run.get("pass_fail"),
            "fragility_flags": sensitivity_run.get("fragility_flags") or [],
            "gate_results": sensitivity_run.get("gate_results") or {},
            "scenario_count": len(list(scenarios)),
        }
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        return path

    def export_replay_aware_model_summary_xlsx(
        self,
        model: dict[str, object],
        evidence_cells: Iterable[dict[str, object]],
    ) -> Path:
        model_version = str(model["model_version"])
        path = self.settings.exports_dir / f"replay_aware_model_summary_{model_version}.xlsx"
        cells = list(evidence_cells)
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(model.get("metrics") or {}).items()]],
            "Training Replay Runs": [["replay_run_id"], *[[run_id] for run_id in list(model.get("replay_run_ids") or [])]],
            "Evidence Overview": [["metric", "value"], ["evidence_cell_count", len(cells)], ["candidate_outcome_row_count", model.get("candidate_outcome_row_count")]],
            "Top Positive Evidence Cells": self._evidence_rows(sorted(cells, key=lambda cell: float(cell.get("average_r") or 0.0), reverse=True)[:100]),
            "Top Negative Evidence Cells": self._evidence_rows(sorted(cells, key=lambda cell: float(cell.get("average_r") or 0.0))[:100]),
            "Suppression Rules": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(model.get("scoring_config") or {}).items() if str(key).startswith(("minimum", "maximum", "block", "suppress"))]],
            "Scoring Config": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(model.get("scoring_config") or {}).items()]],
            "Activation Criteria": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(model.get("activation_criteria") or {}).items()]],
            "Warnings": [["warning"], *[[warning] for warning in list(model.get("warnings") or [])]],
            "Provenance": [
                ["key", "value"],
                ["model_version", model_version],
                ["model_type", model.get("model_type")],
                ["replay_config_hashes", self._cell_value(model.get("replay_config_hashes") or [])],
                ["input_fingerprints", self._cell_value(model.get("input_fingerprints") or [])],
                ["code_version", model.get("code_version")],
            ],
        }
        self._write_workbook(path, sheets)
        return path

    def export_evidence_cells_csv(self, model_version: str, evidence_cells: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"evidence_cells_{model_version}.csv"
        rows = [self._evidence_row(cell) for cell in evidence_cells]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=EVIDENCE_CELL_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_evidence_cells_xlsx(self, model_version: str, evidence_cells: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"evidence_cells_{model_version}.xlsx"
        cells = list(evidence_cells)
        sheets = {
            "Evidence Cells": self._evidence_rows(cells),
            "By Symbol": self._evidence_dimension_sheet(cells, "symbol"),
            "By Setup": self._evidence_dimension_sheet(cells, "setup_type"),
            "By Regime": self._evidence_dimension_sheet(cells, "market_regime"),
            "By Time Bucket": self._evidence_dimension_sheet(cells, "time_bucket"),
            "Fragility Flags": self._flag_sheet(cells, "fragility_flags"),
            "Divergence Flags": self._metric_flag_sheet(cells, "label_vs_replay_divergence_flags"),
        }
        self._write_workbook(path, sheets)
        return path

    def export_score_audits_csv(self, model_version: str, audits: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"score_audits_{model_version}.csv"
        rows = [self._score_audit_row(audit) for audit in audits]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=SCORE_AUDIT_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_score_audits_xlsx(self, model_version: str, audits: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"score_audits_{model_version}.xlsx"
        rows = [self._score_audit_row(audit) for audit in audits]
        self._write_workbook(
            path,
            {"Score Audits": [SCORE_AUDIT_COLUMNS, *[[row.get(column) for column in SCORE_AUDIT_COLUMNS] for row in rows]]},
        )
        return path

    def export_replay_aware_validation_xlsx(self, report: dict[str, object]) -> Path:
        report_id = str(report.get("report_id") or "latest")
        path = self.settings.exports_dir / f"replay_aware_validation_{report_id}.xlsx"
        summary = dict(report.get("summary") or {})
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in summary.items() if not isinstance(value, (dict, list))]],
            "Walk Forward Windows": self._series_sheet(report.get("windows")),
            "Selected Trades": self._series_sheet(report.get("selected_trades")),
            "Suppressed Candidates": self._series_sheet(report.get("suppressed_candidates")),
            "Per Symbol": self._metric_sheet(report.get("per_symbol")),
            "Per Setup": self._metric_sheet(report.get("per_setup")),
            "Per Regime": self._metric_sheet(report.get("per_regime")),
            "Per Time Bucket": self._metric_sheet(report.get("per_time_bucket")),
            "Sensitivity": self._metric_sheet(report.get("sensitivity_summary")),
            "Drawdown": self._series_sheet(summary.get("drawdown_series")),
            "Rejection Reasons": [["reason"], *[[reason] for reason in list(report.get("rejection_reasons") or [])]],
            "Config": [["key", "value"], ["validation_mode", report.get("validation_mode")], ["model_version", report.get("model_version")]],
        }
        self._write_workbook(path, sheets)
        return path

    def export_calibration_audit_xlsx(self, audit: dict[str, object], bins: Iterable[dict[str, object]]) -> Path:
        audit_id = str(audit.get("calibration_audit_id") or "latest")
        path = self.settings.exports_dir / f"calibration_audit_{audit_id}.xlsx"
        stability = dict(audit.get("stability_metrics") or {})
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in audit.items() if key not in {"score_bins", "grade_bins", "action_bins", "stability_metrics"} and not isinstance(value, (dict, list))]],
            "Score Bins": self._dict_table(audit.get("score_bins")),
            "Grade Bins": self._dict_table(audit.get("grade_bins")),
            "Action Bins": self._dict_table(audit.get("action_bins")),
            "Stability By Symbol": self._dict_table(stability.get("by_symbol")),
            "Stability By Setup": self._dict_table(stability.get("by_setup")),
            "Stability By Regime": self._dict_table(stability.get("by_regime")),
            "Stability By Time Bucket": self._dict_table(stability.get("by_time_bucket")),
            "Warnings": [["warning"], *[[warning] for warning in list(audit.get("calibration_warnings") or audit.get("warnings") or [])]],
            "Rejection Reasons": [["reason"], *[[reason] for reason in list(audit.get("rejection_reasons") or [])]],
            "Config": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(audit.get("config") or {}).items()]],
            "Provenance": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(audit.get("provenance") or {}).items()]],
            "Persisted Bins": self._dict_table(list(bins)),
        }
        self._write_workbook(path, sheets)
        return path

    def export_calibration_bins_csv(self, calibration_audit_id: str, bins: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"calibration_bins_{calibration_audit_id}.csv"
        rows = [self._calibration_bin_row(calibration_audit_id, row) for row in bins]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=CALIBRATION_BIN_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_calibration_bins_xlsx(self, calibration_audit_id: str, bins: Iterable[dict[str, object]]) -> Path:
        path = self.settings.exports_dir / f"calibration_bins_{calibration_audit_id}.xlsx"
        rows = [self._calibration_bin_row(calibration_audit_id, row) for row in bins]
        self._write_workbook(
            path,
            {"Calibration Bins": [CALIBRATION_BIN_COLUMNS, *[[row.get(column) for column in CALIBRATION_BIN_COLUMNS] for row in rows]]},
        )
        return path

    def export_calibration_metrics_json(self, audit: dict[str, object]) -> Path:
        audit_id = str(audit.get("calibration_audit_id") or "latest")
        path = self.settings.exports_dir / f"calibration_metrics_{audit_id}.json"
        path.write_text(json.dumps(audit, indent=2, default=str, allow_nan=False), encoding="utf-8")
        return path

    def export_model_comparison_xlsx(self, comparison: dict[str, object]) -> Path:
        comparison_id = str(comparison.get("comparison_id") or "latest")
        path = self.settings.exports_dir / f"model_comparison_{comparison_id}.xlsx"
        ranking = list(comparison.get("robustness_ranking") or [])
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(comparison.get("summary") or {}).items()]],
            "Models": [MODEL_COMPARISON_COLUMNS, *[[row.get(column) for column in MODEL_COMPARISON_COLUMNS] for row in ranking]],
            "Validation Metrics": self._dict_table(comparison.get("validation_reports")),
            "Replay Metrics": self._dict_table(comparison.get("replay_run_ids")),
            "Sensitivity Metrics": [["status", "message"], ["not_applicable", "Sensitivity details are included only when linked to validation artifacts."]],
            "Calibration Metrics": self._dict_table(comparison.get("calibration_audits")),
            "Warnings": [["warning"], *[[warning] for warning in list(comparison.get("warnings") or [])]],
            "Config": [["key", "value"], ["comparison_type", comparison.get("comparison_type")], ["diagnostic_only", True]],
        }
        self._write_workbook(path, sheets)
        return path

    def export_replay_window_set_xlsx(
        self,
        window_set: dict[str, object],
        results: Iterable[dict[str, object]],
    ) -> Path:
        window_set_id = str(window_set.get("window_set_id") or "latest")
        path = self.settings.exports_dir / f"replay_window_set_{window_set_id}.xlsx"
        result_rows = [self._replay_window_result_row(result) for result in results]
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(window_set.get("summary") or {}).items()]],
            "Generated Windows": self._dict_table(window_set.get("generated_windows")),
            "Window Results": [REPLAY_WINDOW_RESULT_COLUMNS, *[[row.get(column) for column in REPLAY_WINDOW_RESULT_COLUMNS] for row in result_rows]],
            "Replay Runs": self._dict_table([run_id for row in result_rows for run_id in list(row.get("replay_run_ids") or [])]),
            "Warnings": [["warning"], *[[warning] for warning in list(window_set.get("warnings") or [])]],
            "Config": [
                ["key", "value"],
                ["window_set_id", window_set_id],
                ["window_mode", window_set.get("window_mode")],
                ["symbols", self._cell_value(window_set.get("symbols") or [])],
                ["intervals", self._cell_value(window_set.get("intervals") or [])],
                ["diagnostic_only", True],
            ],
        }
        self._write_workbook(path, sheets)
        return path

    def export_calibration_drift_xlsx(
        self,
        report: dict[str, object],
        windows: Iterable[dict[str, object]],
    ) -> Path:
        drift_report_id = str(report.get("drift_report_id") or "latest")
        path = self.settings.exports_dir / f"calibration_drift_{drift_report_id}.xlsx"
        window_rows = [self._drift_window_row(drift_report_id, row) for row in windows]
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(report.get("summary") or {}).items()]],
            "Drift Flags": [["flag"], *[[flag] for flag in list(report.get("drift_flags") or [])]],
            "Window Metrics": [DRIFT_WINDOW_COLUMNS, *[[row.get(column) for column in DRIFT_WINDOW_COLUMNS] for row in window_rows]],
            "Score Bin Drift": self._metric_sheet(report.get("score_bin_drift")),
            "Grade Bin Drift": self._metric_sheet(report.get("grade_bin_drift")),
            "Action Bin Drift": self._metric_sheet(report.get("action_bin_drift")),
            "Stability": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(report.get("stability_metrics") or {}).items()]],
            "Warnings": [["warning"], *[[warning] for warning in list(report.get("warnings") or [])]],
            "Config": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(report.get("config") or {}).items()]],
        }
        self._write_workbook(path, sheets)
        return path

    def export_calibration_drift_json(self, report: dict[str, object]) -> Path:
        drift_report_id = str(report.get("drift_report_id") or "latest")
        path = self.settings.exports_dir / f"calibration_drift_{drift_report_id}.json"
        path.write_text(json.dumps(report, indent=2, default=str, allow_nan=False), encoding="utf-8")
        return path

    def export_calibration_drift_windows_csv(
        self,
        drift_report_id: str,
        windows: Iterable[dict[str, object]],
    ) -> Path:
        path = self.settings.exports_dir / f"calibration_drift_windows_{drift_report_id}.csv"
        rows = [self._drift_window_row(drift_report_id, row) for row in windows]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=DRIFT_WINDOW_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        return path

    def export_calibration_drift_windows_xlsx(
        self,
        drift_report_id: str,
        windows: Iterable[dict[str, object]],
    ) -> Path:
        path = self.settings.exports_dir / f"calibration_drift_windows_{drift_report_id}.xlsx"
        rows = [self._drift_window_row(drift_report_id, row) for row in windows]
        self._write_workbook(
            path,
            {"Drift Windows": [DRIFT_WINDOW_COLUMNS, *[[row.get(column) for column in DRIFT_WINDOW_COLUMNS] for row in rows]]},
        )
        return path

    def export_model_review_xlsx(self, report: dict[str, object]) -> Path:
        review_report_id = str(report.get("review_report_id") or "latest")
        path = self.settings.exports_dir / f"model_review_{review_report_id}.xlsx"
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(report.get("summary") or {}).items()]],
            "Readiness": [
                ["key", "value"],
                ["readiness_status", report.get("readiness_status")],
                ["model_version", report.get("model_version")],
                ["window_set_id", report.get("window_set_id")],
                ["model_activation_unchanged", True],
            ],
            "Readiness Reasons": [["reason"], *[[reason] for reason in list(report.get("readiness_reasons") or [])]],
            "Unresolved Warnings": [["warning"], *[[warning] for warning in list(report.get("unresolved_warnings") or [])]],
            "Validation Reports": self._dict_table(report.get("validation_reports")),
            "Calibration Audits": self._dict_table(report.get("calibration_audits")),
            "Drift Reports": self._dict_table(report.get("drift_reports")),
            "Model Summary": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(report.get("model_summary") or {}).items()]],
        }
        self._write_workbook(path, sheets)
        return path

    def export_model_review_json(self, report: dict[str, object]) -> Path:
        review_report_id = str(report.get("review_report_id") or "latest")
        path = self.settings.exports_dir / f"model_review_{review_report_id}.json"
        path.write_text(json.dumps(report, indent=2, default=str, allow_nan=False), encoding="utf-8")
        return path

    def export_research_cycle_xlsx(
        self,
        cycle: dict[str, object],
        artifacts: Iterable[dict[str, object]],
    ) -> Path:
        research_cycle_id = str(cycle.get("research_cycle_id") or "latest")
        path = self.settings.exports_dir / f"research_cycle_{research_cycle_id}.xlsx"
        artifact_rows = list(artifacts)
        sheets = {
            "Summary": [["metric", "value"], *[[key, self._cell_value(value)] for key, value in dict(cycle.get("summary") or {}).items()]],
            "Cycle Config": [["key", "value"], *[[key, self._cell_value(value)] for key, value in dict(cycle.get("config") or {}).items()]],
            "Data Quality": self._artifact_payload_table(artifact_rows, "data_quality"),
            "Stale Windows": self._metric_sheet(cycle.get("stale_window_status")),
            "Replay Windows": self._artifact_payload_table(artifact_rows, "replay_window_set"),
            "Counterfactual Replay": self._id_sheet(cycle.get("counterfactual_replay_run_ids")),
            "Portfolio Replay": self._id_sheet(cycle.get("portfolio_replay_run_ids")),
            "Sensitivity": self._id_sheet(cycle.get("sensitivity_run_ids")),
            "Calibration": self._id_sheet(cycle.get("calibration_audit_ids")),
            "Drift": self._id_sheet(cycle.get("drift_report_ids")),
            "Model Review": self._id_sheet(cycle.get("model_review_report_ids")),
            "Champion vs Challenger": self._id_sheet(cycle.get("comparison_ids")),
            "Proposal": self._id_sheet(cycle.get("proposal_ids")),
            "Warnings": [["warning"], *[[warning] for warning in list(cycle.get("warnings") or [])]],
            "Artifacts": self._dict_table(artifact_rows),
            "Provenance": [
                ["key", "value"],
                ["research_cycle_id", research_cycle_id],
                ["config_hash", cycle.get("config_hash")],
                ["input_fingerprint", cycle.get("input_fingerprint")],
                ["git_commit", cycle.get("git_commit")],
                ["database_revision", cycle.get("database_revision")],
                ["persistence_backend", cycle.get("persistence_backend")],
                ["model_activation_unchanged", True],
            ],
        }
        self._write_workbook(path, sheets)
        return path

    def export_research_cycle_json(
        self,
        cycle: dict[str, object],
        artifacts: Iterable[dict[str, object]],
    ) -> Path:
        research_cycle_id = str(cycle.get("research_cycle_id") or "latest")
        path = self.settings.exports_dir / f"research_cycle_{research_cycle_id}.json"
        path.write_text(
            json.dumps({"cycle": cycle, "artifacts": list(artifacts)}, indent=2, default=str, allow_nan=False),
            encoding="utf-8",
        )
        return path

    def export_model_proposal_xlsx(
        self,
        proposal: dict[str, object],
        ledger: Iterable[dict[str, object]],
    ) -> Path:
        proposal_id = str(proposal.get("proposal_id") or "latest")
        path = self.settings.exports_dir / f"model_proposal_{proposal_id}.xlsx"
        sheets = {
            "Summary": [["key", "value"], ["proposal_id", proposal_id], ["status", proposal.get("status")], ["approval_required", proposal.get("approval_required")]],
            "Recommended Action": [["key", "value"], ["recommended_action", proposal.get("recommended_action")]],
            "Readiness": [["key", "value"], ["readiness_status", proposal.get("readiness_status")]],
            "Champion": [["key", "value"], ["model_version", proposal.get("champion_model_version")], *[[key, self._cell_value(value)] for key, value in dict(proposal.get("champion_metrics") or {}).items()]],
            "Challenger": [["key", "value"], ["model_version", proposal.get("challenger_model_version")], *[[key, self._cell_value(value)] for key, value in dict(proposal.get("challenger_metrics") or {}).items()]],
            "Delta Metrics": self._metric_sheet(proposal.get("delta_metrics")),
            "Gates": self._metric_sheet(proposal.get("pass_fail_gates")),
            "Rejection Reasons": [["reason"], *[[reason] for reason in list(proposal.get("rejection_reasons") or [])]],
            "Approval History": self._dict_table(list(ledger)),
            "Artifacts": [
                ["artifact_type", "ids"],
                ["validation_report_ids", self._cell_value(proposal.get("validation_report_ids") or [])],
                ["calibration_audit_ids", self._cell_value(proposal.get("calibration_audit_ids") or [])],
                ["drift_report_ids", self._cell_value(proposal.get("drift_report_ids") or [])],
                ["model_review_report_ids", self._cell_value(proposal.get("model_review_report_ids") or [])],
                ["comparison_ids", self._cell_value(proposal.get("comparison_ids") or [])],
                ["replay_run_ids", self._cell_value(proposal.get("replay_run_ids") or [])],
                ["window_set_ids", self._cell_value(proposal.get("window_set_ids") or [])],
            ],
            "Provenance": [
                ["key", "value"],
                ["research_cycle_id", proposal.get("research_cycle_id")],
                ["created_at", proposal.get("created_at")],
                ["updated_at", proposal.get("updated_at")],
                ["activation_id", proposal.get("activation_id")],
                ["explicit_activation_required", True],
            ],
        }
        self._write_workbook(path, sheets)
        return path

    def export_model_proposal_json(self, proposal: dict[str, object], ledger: Iterable[dict[str, object]]) -> Path:
        proposal_id = str(proposal.get("proposal_id") or "latest")
        path = self.settings.exports_dir / f"model_proposal_{proposal_id}.json"
        path.write_text(
            json.dumps({"proposal": proposal, "ledger": list(ledger)}, indent=2, default=str, allow_nan=False),
            encoding="utf-8",
        )
        return path

    def export_champion_challenger_comparison_xlsx(self, comparison: dict[str, object]) -> Path:
        comparison_id = str(comparison.get("comparison_id") or "latest")
        path = self.settings.exports_dir / f"champion_challenger_comparison_{comparison_id}.xlsx"
        sheets = {
            "Summary": [
                ["key", "value"],
                ["comparison_id", comparison_id],
                ["recommended_action", comparison.get("recommended_action")],
                ["readiness_status", comparison.get("readiness_status")],
                ["diagnostic_only", True],
            ],
            "Champion": [["key", "value"], ["model_version", comparison.get("champion_model_version")], *[[key, self._cell_value(value)] for key, value in dict(comparison.get("champion_metrics") or {}).items()]],
            "Challenger": [["key", "value"], ["model_version", comparison.get("challenger_model_version")], *[[key, self._cell_value(value)] for key, value in dict(comparison.get("challenger_metrics") or {}).items()]],
            "Delta Metrics": self._metric_sheet(comparison.get("delta_metrics")),
            "Gates": self._metric_sheet(comparison.get("gate_results")),
            "Better Flags": [["flag"], *[[flag] for flag in list(comparison.get("challenger_better_flags") or [])]],
            "Worse Flags": [["flag"], *[[flag] for flag in list(comparison.get("challenger_worse_flags") or [])]],
            "Warnings": [["warning"], *[[warning] for warning in list(comparison.get("warnings") or [])]],
        }
        self._write_workbook(path, sheets)
        return path

    def _signal_row(self, signal: Signal) -> dict[str, object]:
        payload = signal.model_dump(mode="json")
        payload["side"] = signal.side.value
        payload["status"] = signal.status.value
        payload["reasons"] = " | ".join(signal.reasons)
        payload["warnings"] = " | ".join(signal.warnings)
        return {column: payload.get(column) for column in SIGNAL_COLUMNS}

    def _replay_trade_row(self, trade: dict[str, object]) -> dict[str, object]:
        metadata = dict(trade.get("metadata") or {})
        return {column: trade.get(column, metadata.get(column)) for column in REPLAY_TRADE_COLUMNS}

    def _sensitivity_scenario_row(self, scenario: dict[str, object], sensitivity_run_id: str) -> dict[str, object]:
        metrics = dict(scenario.get("summary_metrics") or {})
        return {
            "scenario_id": scenario.get("scenario_id"),
            "sensitivity_run_id": scenario.get("sensitivity_run_id") or sensitivity_run_id,
            "replay_run_id": scenario.get("replay_run_id"),
            "slippage_bps": scenario.get("slippage_bps"),
            "spread_bps": scenario.get("spread_bps"),
            "intrabar_path_policy": scenario.get("intrabar_path_policy"),
            "same_bar_stop_target_policy": scenario.get("same_bar_stop_target_policy"),
            "pass_fail": scenario.get("pass_fail"),
            "total_trades": metrics.get("total_trades"),
            "average_r": metrics.get("average_r"),
            "profit_factor": metrics.get("profit_factor"),
            "max_drawdown_r": metrics.get("max_drawdown_r"),
            "total_r": metrics.get("total_r"),
            "skip_rate": metrics.get("skip_rate"),
        }

    def _evidence_row(self, cell: dict[str, object]) -> dict[str, object]:
        return {column: cell.get(column) for column in EVIDENCE_CELL_COLUMNS}

    def _evidence_rows(self, cells: list[dict[str, object]]) -> list[list[object]]:
        return [EVIDENCE_CELL_COLUMNS, *[[self._evidence_row(cell).get(column) for column in EVIDENCE_CELL_COLUMNS] for cell in cells]]

    def _evidence_dimension_sheet(self, cells: list[dict[str, object]], dimension: str) -> list[list[object]]:
        rows: list[list[object]] = [["dimension", "cell_count", "observed_outcome_count", "average_r"]]
        buckets: dict[str, list[dict[str, object]]] = {}
        for cell in cells:
            dims = dict(cell.get("dimensions") or {})
            value = str(dims.get(dimension) or "unknown")
            buckets.setdefault(value, []).append(cell)
        for value, bucket in sorted(buckets.items()):
            observed = sum(int(cell.get("observed_outcome_count") or 0) for cell in bucket)
            weighted_r = sum(float(cell.get("average_r") or 0.0) * int(cell.get("observed_outcome_count") or 0) for cell in bucket)
            rows.append([value, len(bucket), observed, weighted_r / observed if observed else 0.0])
        return rows

    def _flag_sheet(self, cells: list[dict[str, object]], field: str) -> list[list[object]]:
        rows: list[list[object]] = [["flag", "cell_key"]]
        for cell in cells:
            for flag in list(cell.get(field) or []):
                rows.append([flag, cell.get("cell_key")])
        return rows

    def _metric_flag_sheet(self, cells: list[dict[str, object]], field: str) -> list[list[object]]:
        rows: list[list[object]] = [["flag", "cell_key"]]
        for cell in cells:
            metrics = dict(cell.get("metrics") or {})
            for flag in list(metrics.get(field) or []):
                rows.append([flag, cell.get("cell_key")])
        return rows

    def _score_audit_row(self, audit: dict[str, object]) -> dict[str, object]:
        return {
            "score_id": audit.get("score_id"),
            "model_version": audit.get("model_version"),
            "candidate_id": audit.get("candidate_id"),
            "symbol": audit.get("symbol"),
            "interval": audit.get("interval"),
            "timestamp_utc": audit.get("timestamp_utc"),
            "side": audit.get("side"),
            "setup_type": audit.get("setup_type"),
            "signal_quality_score": audit.get("signal_quality_score"),
            "grade": audit.get("grade"),
            "action": audit.get("action"),
            "expected_r_estimate": audit.get("expected_r_estimate"),
            "suppression_reasons": " | ".join(str(value) for value in list(audit.get("suppression_reasons") or [])),
            "warnings": " | ".join(str(value) for value in list(audit.get("warnings") or [])),
        }

    def _calibration_bin_row(self, calibration_audit_id: str, bin_row: dict[str, object]) -> dict[str, object]:
        metrics = dict(bin_row.get("metrics") or bin_row)
        row = {
            "calibration_audit_id": calibration_audit_id,
            "bin_type": bin_row.get("bin_type"),
            "bin_key": bin_row.get("bin_key"),
            "sample_size": bin_row.get("sample_size"),
            "observed_average_r": bin_row.get("observed_average_r"),
            "observed_win_rate": bin_row.get("observed_win_rate"),
            "profit_factor": bin_row.get("profit_factor"),
            "max_drawdown_r": bin_row.get("max_drawdown_r"),
            "same_bar_ambiguity_rate": metrics.get("same_bar_ambiguity_rate"),
            "fragility_flag_rate": metrics.get("fragility_flag_rate"),
        }
        return {column: row.get(column) for column in CALIBRATION_BIN_COLUMNS}

    def _replay_window_result_row(self, result: dict[str, object]) -> dict[str, object]:
        metrics = dict(result.get("metrics") or {})
        return {
            "window_result_id": result.get("window_result_id"),
            "window_set_id": result.get("window_set_id"),
            "window_index": result.get("window_index"),
            "replay_start": result.get("replay_start"),
            "replay_end": result.get("replay_end"),
            "counterfactual_replay_run_id": result.get("counterfactual_replay_run_id"),
            "portfolio_replay_run_id": result.get("portfolio_replay_run_id"),
            "status": result.get("status"),
            "total_trades": metrics.get("total_trades"),
            "average_r": metrics.get("average_r"),
            "profit_factor": metrics.get("profit_factor"),
            "max_drawdown_r": metrics.get("max_drawdown_r"),
            "warning_count": len(list(result.get("warnings") or [])),
            "replay_run_ids": list(result.get("replay_run_ids") or []),
        }

    def _drift_window_row(self, drift_report_id: str, window: dict[str, object]) -> dict[str, object]:
        metrics = dict(window.get("metrics") or {})
        flags = list(window.get("flags") or [])
        return {
            "drift_report_id": drift_report_id,
            "window_result_id": window.get("window_result_id"),
            "window_index": window.get("window_index"),
            "severity": window.get("severity"),
            "rank_correlation_score": metrics.get("rank_correlation_score"),
            "monotonicity_pass": metrics.get("monotonicity_pass"),
            "high_grade_average_r": metrics.get("high_grade_average_r"),
            "take_minus_watch_average_r": metrics.get("take_minus_watch_average_r"),
            "flags": " | ".join(str(flag) for flag in flags),
        }

    def _replay_workbook_sheets(
        self,
        replay_run: dict[str, object],
        metrics: dict[str, object],
        trade_rows: list[dict[str, object]],
        skipped_rows: list[dict[str, object]],
    ) -> dict[str, list[list[object]]]:
        drawdown_payload = metrics.get("drawdown_series")
        drawdown_values = drawdown_payload if isinstance(drawdown_payload, list) else []
        config_payload = replay_run.get("config")
        config = config_payload if isinstance(config_payload, dict) else {}
        warnings_payload = replay_run.get("warnings")
        warnings = warnings_payload if isinstance(warnings_payload, list) else []
        return {
            "Summary": [["metric", "value"], *[[key, json.dumps(value, default=str)] for key, value in metrics.items() if not isinstance(value, (dict, list))]],
            "Trades": [REPLAY_TRADE_COLUMNS, *[[row.get(column) for column in REPLAY_TRADE_COLUMNS] for row in trade_rows if row.get("status") == "TAKEN"]],
            "Skipped Candidates": [REPLAY_TRADE_COLUMNS, *[[row.get(column) for column in REPLAY_TRADE_COLUMNS] for row in skipped_rows]],
            "Per Symbol": self._metric_sheet(metrics.get("per_symbol_metrics")),
            "Per Setup": self._metric_sheet(metrics.get("per_setup_metrics")),
            "Per Regime": self._metric_sheet(metrics.get("per_regime_metrics")),
            "Per Time Bucket": self._metric_sheet(metrics.get("per_time_bucket_metrics")),
            "Daily R": self._series_sheet(metrics.get("daily_r_series")),
            "Drawdown": [["index", "drawdown_r"], *[[index, value] for index, value in enumerate(drawdown_values, start=1)]],
            "Config": [["key", "value"], *[[key, json.dumps(value, default=str)] for key, value in config.items()]],
            "Warnings": [["warning"], *[[warning] for warning in warnings]],
        }

    def _sensitivity_workbook_sheets(
        self,
        sensitivity_run: dict[str, object],
        scenario_rows: list[dict[str, object]],
    ) -> dict[str, list[list[object]]]:
        worst = dict(sensitivity_run.get("worst_case") or {})
        median_case = dict(sensitivity_run.get("median_case") or {})
        best = dict(sensitivity_run.get("best_case") or {})
        fragility_payload = sensitivity_run.get("fragility_flags")
        fragility_flags = fragility_payload if isinstance(fragility_payload, list) else []
        return {
            "Summary": [
                ["metric", "value"],
                ["sensitivity_run_id", sensitivity_run.get("sensitivity_run_id")],
                ["replay_run_id", sensitivity_run.get("replay_run_id")],
                ["scenario_count", sensitivity_run.get("scenario_count")],
                ["robustness_score", sensitivity_run.get("robustness_score")],
                ["pass_fail", sensitivity_run.get("pass_fail")],
            ],
            "Scenario Metrics": [
                SENSITIVITY_SCENARIO_COLUMNS,
                *[[row.get(column) for column in SENSITIVITY_SCENARIO_COLUMNS] for row in scenario_rows],
            ],
            "Worst Case": self._scenario_detail_sheet(worst),
            "Median Case": self._scenario_detail_sheet(median_case),
            "Best Case": self._scenario_detail_sheet(best),
            "Fragility Flags": [["flag"], *[[flag] for flag in fragility_flags]],
            "Gate Results": [["gate", "passed"], *[[key, value] for key, value in dict(sensitivity_run.get("gate_results") or {}).items()]],
            "Config": [["key", "value"], *[[key, json.dumps(value, default=str)] for key, value in dict(sensitivity_run.get("config") or {}).items()]],
            "Warnings": [["warning"], ["audit", "Replay sensitivity is a stress test, not a profitability claim."]],
        }

    def _scenario_detail_sheet(self, scenario: dict[str, object]) -> list[list[object]]:
        rows: list[list[object]] = [["section", "key", "value"]]
        for key, value in scenario.items():
            if key in {"summary_metrics", "gate_results"} and isinstance(value, dict):
                rows.extend([[key, metric_key, self._cell_value(metric_value)] for metric_key, metric_value in value.items()])
            else:
                rows.append(["scenario", key, json.dumps(value, default=str)])
        return rows

    def _cell_value(self, value: object) -> object:
        if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
            return None
        if isinstance(value, datetime):
            if value.tzinfo is not None:
                value = value.astimezone(UTC).replace(tzinfo=None)
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (dict, list, tuple, set)):
            return json.dumps(value, default=str)
        return value

    def _metric_sheet(self, payload: object) -> list[list[object]]:
        rows: list[list[object]] = [["group", "metric", "value"]]
        if isinstance(payload, dict):
            for group, metrics in payload.items():
                if isinstance(metrics, dict):
                    rows.extend([[group, key, value] for key, value in metrics.items()])
        return rows

    def _series_sheet(self, payload: object) -> list[list[object]]:
        rows: list[list[object]] = [["key", "value"]]
        if isinstance(payload, list):
            for item in payload:
                if isinstance(item, dict):
                    value = item.get("r") if "r" in item else item.get("value")
                    rows.append([item.get("date") or item.get("key"), value])
                else:
                    rows.append([len(rows), item])
        return rows

    def _dict_table(self, payload: object) -> list[list[object]]:
        rows = list(payload) if isinstance(payload, list) else []
        if not rows:
            return [["status", "message"], ["empty", "No rows available."]]
        if not all(isinstance(row, dict) for row in rows):
            return [["value"], *[[self._cell_value(row)] for row in rows]]
        columns = sorted({str(key) for row in rows for key in row.keys()})
        return [columns, *[[self._cell_value(row.get(column)) for column in columns] for row in rows]]

    def _write_sheet(self, sheet, columns: list[str], rows: list[dict[str, object]]) -> None:
        sheet.append(columns)
        for row in rows:
            sheet.append([self._cell_value(row.get(column)) for column in columns])

    def _id_sheet(self, payload: object) -> list[list[object]]:
        values = list(payload) if isinstance(payload, list) else []
        return [["id"], *[[self._cell_value(value)] for value in values]] if values else [["status", "message"], ["empty", "No IDs available."]]

    def _artifact_payload_table(self, artifacts: list[dict[str, object]], artifact_type: str) -> list[list[object]]:
        matches = [artifact for artifact in artifacts if artifact.get("artifact_type") == artifact_type]
        payloads = [artifact.get("payload") or artifact for artifact in matches]
        return self._dict_table(payloads)

    def _write_workbook(self, path: Path, sheets: dict[str, list[list[object]]]) -> None:
        if Workbook is None:
            self._write_minimal_xlsx(path, sheets)
            return
        workbook = Workbook()
        for index, (sheet_name, rows) in enumerate(sheets.items()):
            sheet = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
            sheet.title = sheet_name
            for row in rows:
                sheet.append([self._cell_value(value) for value in row])
        workbook.save(path)

    def _write_minimal_xlsx(self, path: Path, sheets: dict[str, list[list[object]]]) -> None:
        with ZipFile(path, "w", ZIP_DEFLATED) as archive:
            archive.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
"""
                + "".join(
                    f'  <Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>\n'
                    for index in range(1, len(sheets) + 1)
                )
                + "</Types>",
            )
            archive.writestr(
                "_rels/.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
            )
            archive.writestr(
                "xl/workbook.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
"""
                + "".join(
                    f'    <sheet name="{escape(name[:31])}" sheetId="{index}" r:id="rId{index}"/>\n'
                    for index, name in enumerate(sheets, start=1)
                )
                + "  </sheets>\n</workbook>",
            )
            archive.writestr(
                "xl/_rels/workbook.xml.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
"""
                + "".join(
                    f'  <Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>\n'
                    for index in range(1, len(sheets) + 1)
                )
                + "</Relationships>",
            )
            for index, rows in enumerate(sheets.values(), start=1):
                archive.writestr(f"xl/worksheets/sheet{index}.xml", self._sheet_xml(rows))

    def _sheet_xml(self, rows: list[list[object]]) -> str:
        xml_rows = []
        for row_index, row in enumerate(rows, start=1):
            cells = []
            for column_index, value in enumerate(row, start=1):
                cell_ref = f"{self._column_name(column_index)}{row_index}"
                text = "" if value is None else str(value)
                cells.append(f'<c r="{cell_ref}" t="inlineStr"><is><t>{escape(text)}</t></is></c>')
            xml_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<sheetData>{"".join(xml_rows)}</sheetData>'
            "</worksheet>"
        )

    def _column_name(self, index: int) -> str:
        name = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name
        return name
